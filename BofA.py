import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

class Bond:
    def __init__(self, bond_id, coupon, frequency, months_since_coupon):
        self.id = bond_id
        self.coupon = float(coupon)
        self.frequency = int(float(frequency))
        self.months_since_coupon = int(float(months_since_coupon))
        # Accrued Interest per 100 face value
        # AI = (Coupon / Frequency) * (MonthsSinceCoupon / (12 / Frequency))
        #    = Coupon * MonthsSinceCoupon / 12  (simplified, expressed as decimal)
        # Multiply by 100 to get price-equivalent (per 100 face)
        self.accrued_interest = (self.coupon * self.months_since_coupon / 12) * 100

    def __repr__(self):
        return f"Bond({self.id}, coupon={self.coupon}, freq={self.frequency}, AI={self.accrued_interest:.4f})"


class Event:
    def __init__(self, event_id, desk, trader, bond_id, buy_sell, quantity, clean_price):
        self.id = int(float(event_id))
        self.desk = desk
        self.trader = trader
        self.bond_id = bond_id
        self.buy_sell = buy_sell
        self.quantity = int(float(quantity))
        self.clean_price = float(clean_price)

    @property
    def signed_quantity(self):
        """Positive for BUY, negative for SELL."""
        return self.quantity if self.buy_sell == "BUY" else -self.quantity


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_bonds(path="bonds.xlsx"):
    bonds = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        # Skip empty / blank rows until we find the real header
        if not any(c for c in row if c is not None):
            continue
        stripped = [str(c).strip() if c is not None else "" for c in row]
        if header is None:
            header = stripped
            continue
        data = dict(zip(header, stripped))
        if not data.get("BondID"):
            continue
        b = Bond(data["BondID"], data["Coupon"], data["Frequency"], data["MonthsSinceCoupon"])
        bonds[b.id] = b
    return bonds


def load_events(path="events.xlsx"):
    events = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        if not any(c for c in row if c is not None):
            continue
        stripped = [str(c).strip() if c is not None else "" for c in row]
        if header is None:
            header = stripped
            continue
        data = dict(zip(header, stripped))
        if not data.get("EventID"):
            continue
        e = Event(
            data["EventID"], data["Desk"], data["Trader"],
            data["BondID"], data["BuySell"], data["Quantity"], data["CleanPrice"],
        )
        events.append(e)
    return events


# ---------------------------------------------------------------------------
# Calculation helpers
# ---------------------------------------------------------------------------

def dirty_price(clean_price, accrued_interest):
    """Dirty Price = Clean Price + Accrued Interest (both per 100 face)."""
    return clean_price + accrued_interest


def trade_pv(event, accrued_interest):
    """PV for a single trade = Dirty Price * signed quantity."""
    dp = dirty_price(event.clean_price, accrued_interest)
    return dp * event.signed_quantity


def compute_book(events, bonds, up_to_event=None):
    """
    Process events (optionally up to a given EventID) and return:
      - positions: {(bond_id, desk, trader): net_qty}
      - pv_book:   {(bond_id, desk, trader): cumulative PV}
      - event_details: list of dicts with per-event calculated fields
    """
    positions = defaultdict(int)
    pv_book = defaultdict(float)
    event_details = []

    for ev in events:
        if up_to_event is not None and ev.id > up_to_event:
            break

        bond = bonds[ev.bond_id]
        ai = bond.accrued_interest
        dp = dirty_price(ev.clean_price, ai)
        pv = dp * ev.signed_quantity

        key = (ev.bond_id, ev.desk, ev.trader)
        positions[key] += ev.signed_quantity
        pv_book[key] += pv

        event_details.append({
            "EventID": ev.id,
            "Desk": ev.desk,
            "Trader": ev.trader,
            "BondID": ev.bond_id,
            "BuySell": ev.buy_sell,
            "Quantity": ev.quantity,
            "CleanPrice": ev.clean_price,
            "AccruedInterest": ai,
            "DirtyPrice": dp,
            "PV": pv,
            "Position": positions[key],
            "CumulativePV": pv_book[key],
        })

    return positions, pv_book, event_details


# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------

SEPARATOR = "-" * 100

def print_table(headers, rows, col_widths=None):
    """Print a simple formatted table."""
    if not rows:
        print("  (no data)")
        return

    # Auto-compute widths if not given
    if col_widths is None:
        col_widths = []
        for i, h in enumerate(headers):
            max_w = len(str(h))
            for r in rows:
                max_w = max(max_w, len(str(r[i])))
            col_widths.append(max_w + 2)

    fmt = "  ".join(f"{{:<{w}}}" for w in col_widths)
    print(fmt.format(*headers))
    print("  ".join("-" * w for w in col_widths))
    for r in rows:
        print(fmt.format(*r))


def format_num(val, decimals=2):
    return f"{val:,.{decimals}f}"


# ---------------------------------------------------------------------------
# Query handlers
# ---------------------------------------------------------------------------

def query_bond(bond_id, positions, pv_book, bonds):
    """Show info and positions for a specific bond."""
    bond_id = bond_id.upper()
    if bond_id not in bonds:
        print(f"  Bond '{bond_id}' not found.")
        return

    bond = bonds[bond_id]
    print(f"\n  Bond ID             : {bond.id}")
    print(f"  Coupon Rate         : {bond.coupon:.2%}")
    print(f"  Frequency           : {bond.frequency}x/year")
    print(f"  Months Since Coupon : {bond.months_since_coupon}")
    print(f"  Accrued Interest    : {format_num(bond.accrued_interest, 4)}")
    print()

    headers = ["Desk", "Trader", "Position", "PV"]
    rows = []
    total_pos = 0
    total_pv = 0.0
    for (bid, desk, trader), pos in sorted(positions.items()):
        if bid != bond_id:
            continue
        pv = pv_book[(bid, desk, trader)]
        rows.append((desk, trader, pos, format_num(pv)))
        total_pos += pos
        total_pv += pv

    print_table(headers, rows)
    print()
    print(f"  Total Position : {total_pos}")
    print(f"  Total PV       : {format_num(total_pv)}")


def query_desk(desk_name, positions, pv_book, bonds):
    """Show all positions for a desk."""
    desk_name = desk_name.upper()
    headers = ["BondID", "Trader", "Position", "AccruedInt", "PV"]
    rows = []
    total_pv = 0.0

    for (bid, desk, trader), pos in sorted(positions.items()):
        if desk != desk_name:
            continue
        ai = bonds[bid].accrued_interest
        pv = pv_book[(bid, desk, trader)]
        rows.append((bid, trader, pos, format_num(ai, 4), format_num(pv)))
        total_pv += pv

    if not rows:
        print(f"  No positions found for desk '{desk_name}'.")
        return

    print(f"\n  Desk: {desk_name}")
    print()
    print_table(headers, rows)
    print()
    print(f"  Total PV : {format_num(total_pv)}")


def query_trader(trader_name, positions, pv_book, bonds):
    """Show all positions for a trader."""
    trader_name = trader_name.upper()
    headers = ["BondID", "Desk", "Position", "AccruedInt", "PV"]
    rows = []
    total_pv = 0.0

    for (bid, desk, trader), pos in sorted(positions.items()):
        if trader.upper() != trader_name:
            continue
        ai = bonds[bid].accrued_interest
        pv = pv_book[(bid, desk, trader)]
        rows.append((bid, desk, pos, format_num(ai, 4), format_num(pv)))
        total_pv += pv

    if not rows:
        print(f"  No positions found for trader '{trader_name}'.")
        return

    print(f"\n  Trader: {trader_name}")
    print()
    print_table(headers, rows)
    print()
    print(f"  Total PV : {format_num(total_pv)}")


def query_event(event_id, event_details):
    """Show details for a specific event."""
    for ed in event_details:
        if ed["EventID"] == event_id:
            print(f"\n  Event ID        : {ed['EventID']}")
            print(f"  Desk            : {ed['Desk']}")
            print(f"  Trader          : {ed['Trader']}")
            print(f"  Bond ID         : {ed['BondID']}")
            print(f"  Buy/Sell        : {ed['BuySell']}")
            print(f"  Quantity        : {ed['Quantity']}")
            print(f"  Clean Price     : {format_num(ed['CleanPrice'])}")
            print(f"  Accrued Interest: {format_num(ed['AccruedInterest'], 4)}")
            print(f"  Dirty Price     : {format_num(ed['DirtyPrice'])}")
            print(f"  Trade PV        : {format_num(ed['PV'])}")
            print(f"  Running Position: {ed['Position']}")
            print(f"  Cumulative PV   : {format_num(ed['CumulativePV'])}")
            return
    print(f"  Event {event_id} not found.")


def query_events_range(start, end, event_details, bonds):
    """Show a table of events in a range."""
    headers = ["ID", "Desk", "Trader", "Bond", "B/S", "Qty", "CleanPx", "AI", "DirtyPx", "TradePV"]
    rows = []
    for ed in event_details:
        if ed["EventID"] < start or ed["EventID"] > end:
            continue
        rows.append((
            ed["EventID"], ed["Desk"], ed["Trader"], ed["BondID"],
            ed["BuySell"], ed["Quantity"],
            format_num(ed["CleanPrice"]), format_num(ed["AccruedInterest"], 4),
            format_num(ed["DirtyPrice"]), format_num(ed["PV"]),
        ))
    print()
    print_table(headers, rows)


def query_pv_summary(positions, pv_book, bonds):
    """Show PV summary by bond."""
    headers = ["BondID", "Net Position", "Total PV"]
    bond_agg = defaultdict(lambda: [0, 0.0])
    grand_pv = 0.0

    for (bid, desk, trader), pos in positions.items():
        pv = pv_book[(bid, desk, trader)]
        bond_agg[bid][0] += pos
        bond_agg[bid][1] += pv
        grand_pv += pv

    rows = []
    for bid in sorted(bond_agg):
        net_pos, total_pv = bond_agg[bid]
        rows.append((bid, net_pos, format_num(total_pv)))

    print()
    print_table(headers, rows)
    print()
    print(f"  Grand Total PV : {format_num(grand_pv)}")


def query_positions(positions, pv_book):
    """Show all positions."""
    headers = ["BondID", "Desk", "Trader", "Position", "PV"]
    rows = []
    for (bid, desk, trader), pos in sorted(positions.items()):
        pv = pv_book[(bid, desk, trader)]
        rows.append((bid, desk, trader, pos, format_num(pv)))
    print()
    print_table(headers, rows)


def query_after_event(event_id, events, bonds):
    """Recompute the book up to a specific event and show summary."""
    pos, pvb, details = compute_book(events, bonds, up_to_event=event_id)
    print(f"\n  === State after Event {event_id} ===")
    query_pv_summary(pos, pvb, bonds)
    return pos, pvb, details


def last_clean_price_for_bond(bond_id, events, up_to_event=None):
    """Return the clean price from the most recent event for bond_id (up to event N)."""
    last_price = None
    for ev in events:
        if up_to_event is not None and ev.id > up_to_event:
            break
        if ev.bond_id == bond_id:
            last_price = ev.clean_price
    return last_price


def show_bond(bond_id, events, bonds, up_to_event=None):
    """Show net position, last clean price, dirty price and PV for a bond."""
    bond_id = bond_id.upper()
    if bond_id not in bonds:
        print(f"\n  Bond '{bond_id}' not found.")
        return

    label = f"at Event {up_to_event}" if up_to_event is not None else str(up_to_event)
    # compute positions up to the event
    pos, pvb, _ = compute_book(events, bonds, up_to_event=up_to_event)

    net_pos = sum(qty for (bid, _, _), qty in pos.items() if bid == bond_id)
    clean_px = last_clean_price_for_bond(bond_id, events, up_to_event=up_to_event)

    if clean_px is None or net_pos == 0:
        tag = f" at Event {up_to_event}" if up_to_event else ""
        print(f"\n  No data for bond {bond_id}{tag}.")
        return

    ai = bonds[bond_id].accrued_interest
    dp = clean_px + ai
    pv = dp * net_pos

    tag = f" at Event {up_to_event}" if up_to_event is not None else ""
    print(f"\n  Bond {bond_id}{tag}")
    print()
    headers = ["Metric", "Value"]
    rows = [
        ("Position",    net_pos),
        ("Clean Price", format_num(clean_px)),
        ("Dirty Price", format_num(dp)),
        ("PV",          format_num(pv)),
    ]
    print_table(headers, rows, col_widths=[14, 14])


def query_pnl_bond_since(bond_id, event_id, events, bonds, positions, pv_book):
    """P&L for a specific bond from event N to now."""
    bond_id = bond_id.upper()
    if bond_id not in bonds:
        print(f"\n  Bond '{bond_id}' not found.")
        return

    pos_before, pv_before, _ = compute_book(events, bonds, up_to_event=event_id - 1)

    pv_now   = sum(v for (bid, _, _), v in pv_book.items()   if bid == bond_id)
    pv_prior = sum(v for (bid, _, _), v in pv_before.items() if bid == bond_id)
    pnl = pv_now - pv_prior

    print(f"\n  P&L for {bond_id} since Event {event_id}: {format_num(pnl)}")


def query_pnl_since(event_id, events, bonds, positions, pv_book):
    """Show incremental P&L from event_id to the latest event."""
    pos_before, pv_before, _ = compute_book(events, bonds, up_to_event=event_id - 1)

    all_keys = set(pv_before.keys()) | set(pv_book.keys())
    headers = ["BondID", "Desk", "Trader", "PV Before", "PV Now", "P&L"]
    rows = []
    total_pnl = 0.0

    for key in sorted(all_keys):
        bid, desk, trader = key
        before = pv_before[key]
        now = pv_book[key]
        pnl = now - before
        total_pnl += pnl
        rows.append((bid, desk, trader, format_num(before), format_num(now), format_num(pnl)))

    print(f"\n  === P&L since Event {event_id} ===")
    print()
    print_table(headers, rows)
    print()
    print(f"  Total P&L : {format_num(total_pnl)}")


# ---------------------------------------------------------------------------
# Interactive dashboard
# ---------------------------------------------------------------------------

HELP_TEXT = """
  Available commands:
  ─────────────────────────────────────────────────────────────
  show bond <BOND_ID>                 Position, prices & PV for a bond
  show bond <BOND_ID> at event <N>    Bond state at event N
  bond <BOND_ID>                      Positions by desk/trader for a bond
  desk <DESK>                         Positions & PV for a desk (NY/LN/HK)
  trader <TRADER>                     Positions & PV for a trader
  event <ID>                          Details for a single event
  events <FROM> <TO>                  Table of events in a range
  after <EVENT_ID>                    PV summary at state after a specific event
  pv after <EVENT_ID>                 Same as above
  pnl bond <BOND_ID> since event <N>  P&L for one bond since event N
  pnl since <EVENT_ID>                P&L all positions since event N
  positions                           All current positions
  pv                                  PV summary by bond
  help                                Show this help message
  quit / exit                         Exit the dashboard
  ─────────────────────────────────────────────────────────────
"""


def main():
    print("\n  Loading data...")
    bonds = load_bonds()
    events = load_events()
    positions, pv_book, event_details = compute_book(events, bonds)

    print(f"  Loaded {len(bonds)} bonds and {len(events)} events.\n")
    print("  Bond Dashboard — Interactive Query Tool")
    print(HELP_TEXT)

    while True:
        try:
            user_input = input("  >> ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n  Goodbye!")
            break

        if not user_input:
            continue

        parts = user_input.split()
        cmd = parts[0].lower()

        if cmd in ("quit", "exit", "q"):
            print("  Goodbye!")
            break
        elif cmd == "help":
            print(HELP_TEXT)
        elif cmd == "show" and len(parts) >= 3 and parts[1].lower() == "bond":
            bond_id = parts[2]
            # show bond BOND1 at event 50
            if len(parts) >= 6 and parts[3].lower() == "at" and parts[4].lower() == "event":
                try:
                    show_bond(bond_id, events, bonds, up_to_event=int(parts[5]))
                except ValueError:
                    print("  Usage: show bond <BOND_ID> at event <N>")
            else:
                show_bond(bond_id, events, bonds)
        elif cmd == "bond" and len(parts) >= 2:
            query_bond(parts[1], positions, pv_book, bonds)
        elif cmd == "desk" and len(parts) >= 2:
            query_desk(parts[1], positions, pv_book, bonds)
        elif cmd == "trader" and len(parts) >= 2:
            query_trader(parts[1], positions, pv_book, bonds)
        elif cmd == "event" and len(parts) >= 2:
            try:
                eid = int(parts[1])
                query_event(eid, event_details)
            except ValueError:
                print("  Usage: event <ID>  (ID must be a number)")
        elif cmd == "events" and len(parts) >= 3:
            try:
                query_events_range(int(parts[1]), int(parts[2]), event_details, bonds)
            except ValueError:
                print("  Usage: events <FROM> <TO>  (both must be numbers)")
        elif cmd == "after" and len(parts) >= 2:
            try:
                eid = int(parts[1])
                query_after_event(eid, events, bonds)
            except ValueError:
                print("  Usage: after <EVENT_ID>  (ID must be a number)")
        elif cmd == "pv" and len(parts) >= 3 and parts[1].lower() == "after":
            try:
                eid = int(parts[2])
                query_after_event(eid, events, bonds)
            except ValueError:
                print("  Usage: pv after <EVENT_ID>  (ID must be a number)")
        elif cmd in ("pnl", "p&l") and len(parts) >= 5 and parts[1].lower() == "bond" and parts[3].lower() == "since":
            # pnl bond BOND4 since event 56  (parts[4] may be "event" and parts[5] the number)
            try:
                if parts[3].lower() == "since" and parts[4].lower() == "event":
                    eid = int(parts[5])
                else:
                    eid = int(parts[4])
                query_pnl_bond_since(parts[2], eid, events, bonds, positions, pv_book)
            except (ValueError, IndexError):
                print("  Usage: pnl bond <BOND_ID> since event <N>")
        elif cmd in ("pnl", "p&l") and len(parts) >= 3 and parts[1].lower() == "since":
            try:
                eid = int(parts[2])
                query_pnl_since(eid, events, bonds, positions, pv_book)
            except ValueError:
                print("  Usage: pnl since <EVENT_ID>  (ID must be a number)")
        elif cmd in ("positions", "position", "pos"):
            query_positions(positions, pv_book)
        elif cmd == "pv":
            query_pv_summary(positions, pv_book, bonds)
        else:
            print(f"  Unknown command: '{user_input}'. Type 'help' for available commands.")

        print()


if __name__ == "__main__":
    main()