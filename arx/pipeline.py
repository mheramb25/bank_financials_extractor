"""The orchestrator: Stages 1-16, wired together.

This is the *only* place the stage order lives.  ``run.py`` (CLI) and ``app.py``
(Streamlit) are both thin wrappers around :func:`run_batch` -- there is no
duplicated logic between them, and there is no code path that a test cannot
reach without a PDF (``process_document`` takes a :class:`Document`, which the
tests build by hand).

Two-pass scoring, and the reason for it
---------------------------------------
Half the checks in this system are *cross-metric*: you cannot test ``GNPA >=
NNPA`` until you have provisionally chosen both, and you cannot test
``NII = Interest Earned - Interest Expended`` until you have provisionally chosen
all three.  So:

  Pass 1  trial every candidate; score it on provenance alone; take a
          provisional winner per metric, but only trust the ones that already
          clear the write threshold.
  Pass 2  run the banking rules, the formulas, the reverse-validation index, the
          historical and peer comparisons over that provisional set; fold the
          results back in; re-score.

A metric whose provisional value was junk therefore cannot poison another
metric's checks, because it never entered the provisional set.
"""

from __future__ import annotations

import json
import logging
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from tqdm import tqdm

from arx import load_config, load_metrics, metrics_by_key, seed_everything
from arx.confidence import (
    Decision,
    na_reason,
    rejection_reason,
    score_verdict,
    writes_to_excel,
)
from arx.courtroom import TrialContext, hold_trial
from arx.dna import (
    DnaContext,
    dna_score,
    level1_range,
    level2_rules,
    level3_history,
    level4_peers,
    metrics_in_rule,
    reverse_validate,
)
from arx.excel_writer import peer_values, read_history, write_results
from arx.extract import collect_evidence, compile_metrics, generate_candidates
from arx.formulas import (
    check_formulas,
    derivable,
    formula_score,
    sanity_checks,
    trend_check,
)
from arx.models import (
    Candidate,
    CellResult,
    CheckResult,
    Document,
    DocumentResult,
    MetricDef,
    Verdict,
)
from arx.parse import enrich_tables_with_camelot, ocr_pages, parse_pdf
from arx.rank import rank_pages, scanned_pages_in
from arx import load_banking_rules

log = logging.getLogger("arx.pipeline")


# --------------------------------------------------------------------------- #
# Stage 1-5 for one PDF
# --------------------------------------------------------------------------- #


def build_document(
    pdf_path: str | Path,
    cfg: Optional[dict] = None,
    use_cache: bool = True,
) -> tuple[Document, Dict[str, List[Candidate]]]:
    """Parse, rank, OCR-if-needed, camelot-if-needed, and generate candidates."""
    cfg = cfg or load_config()
    doc = parse_pdf(pdf_path, cfg, use_cache=use_cache)

    ranked = rank_pages(doc, cfg)

    # Stage 1, lazily: OCR only the ranked pages that actually look scanned.
    to_ocr = scanned_pages_in(ranked, cfg)
    if to_ocr:
        ocr_pages(doc, to_ocr, cfg)
        ranked = rank_pages(doc, cfg)  # OCR changes section labels and density

    # Camelot is expensive; run it only where the metrics actually live.
    top_pages = [p.number for p in ranked[:80]]
    enrich_tables_with_camelot(doc, top_pages, cfg)

    compiled = compile_metrics()
    pool = generate_candidates(doc, ranked, compiled, cfg)
    pool = collect_evidence(pool, cfg)
    return doc, pool


# --------------------------------------------------------------------------- #
# Stages 6-15 for one document
# --------------------------------------------------------------------------- #


def process_document(
    doc: Document,
    pool: Dict[str, List[Candidate]],
    history: Optional[Dict[str, Dict[str, Dict[str, float]]]] = None,
    category_of: Optional[Dict[str, str]] = None,
    cfg: Optional[dict] = None,
) -> DocumentResult:
    """Turn a candidate pool into decided, audited cells.

    Pure function of its inputs -- no file I/O, no PDF -- which is what makes the
    end-to-end test runnable offline in milliseconds.
    """
    cfg = cfg or load_config()
    history = history or {}
    category_of = category_of or {}
    metrics = load_metrics()
    by_key = metrics_by_key()

    target_fy = doc.fiscal_year or "FY22-23"
    prior_fy = doc.prior_fiscal_year

    result = DocumentResult(
        path=doc.path,
        institution=doc.institution or Path(doc.path).stem,
        category=doc.category or "",
        inst_type=doc.inst_type,
        fiscal_year=target_fy,
        prior_fiscal_year=prior_fy,
        candidate_pool=pool,
    )

    # ---------------- Pass 1: provenance-only trial ------------------------ #
    verdicts: Dict[str, Verdict] = {}
    all_verdicts: Dict[str, List[Verdict]] = {}

    for md in metrics:
        if not md.applies_to(doc.inst_type):
            continue
        cands = pool.get(md.key, [])
        if not cands:
            continue
        ctx = TrialContext(
            target_fy=target_fy,
            prior_fy=prior_fy,
            metric=md,
            inst_type=doc.inst_type,
            cfg=cfg,
        )
        vs = hold_trial(cands, ctx)
        if not vs:
            continue
        all_verdicts[md.key] = vs
        verdicts[md.key] = score_verdict(vs[0], md, cfg)

    write_floor = float(cfg["decision_bands"]["manual_review"])
    provisional: Dict[str, float] = {
        key: v.candidate.value
        for key, v in verdicts.items()
        if v.candidate.value is not None and _is_provisional(v, by_key[key], write_floor)
    }

    # Prior-year column: a CROSS-CHECK for the earlier sheet, never the primary
    # source for it (Stage 2). We keep it only for trend validation.
    prior_values = _prior_year_values(pool, prior_fy, cfg)
    result.prior_year_values = prior_values

    # ---------------- Pass 2: cross-metric checks -------------------------- #
    inst = result.institution
    dna_ctx = DnaContext(
        inst_type=doc.inst_type,
        fiscal_year=target_fy,
        history=history.get(inst, {}),
        peers=peer_values(
            history, category_of, result.category, target_fy, exclude=inst
        ),
        cfg=cfg,
    )

    rule_results = level2_rules(provisional)
    formula_results = check_formulas(provisional, cfg)
    failed_rules_by_metric: Dict[str, List[CheckResult]] = {}
    rules = {r.id: r for r in load_banking_rules()}
    for check in rule_results:
        rule = rules.get(check.name)
        if rule is None:
            continue
        for key in metrics_in_rule(rule):
            failed_rules_by_metric.setdefault(key, []).append(check)

    for key, verdict in verdicts.items():
        md = by_key[key]
        cand = verdict.candidate
        checks: List[CheckResult] = [level1_range(cand.value, md)]
        checks.extend(sanity_checks(cand.value, md, cand.raw_text))
        checks.extend(failed_rules_by_metric.get(key, []))

        h = level3_history(key, cand.value, dna_ctx)
        if h:
            checks.append(h)
        p = level4_peers(key, cand.value, dna_ctx)
        if p:
            checks.append(p)

        # Stage 10: value -> metric.
        rv = reverse_validate(key, cand, pool, cfg)
        if rv:
            verdict.prosecutor_hits.append(rv)
            verdict.court_score = verdict.defence_score - verdict.total_penalty

        tscore, tcheck = trend_check(key, cand.value, prior_values.get(key), cfg)
        if tcheck:
            checks.append(tcheck)

        verdict.checks = checks + [
            c for c in formula_results if key in _formula_metrics(c.name)
        ]
        verdict.dna_score = dna_score(
            [c for c in checks if c.level in ("L1", "L2", "L3", "L4")], cfg
        )
        verdict.formula_score = formula_score(key, formula_results, cfg)
        verdict.trend_score = tscore

        # Numerical-sanity failures are DNA failures too -- they are the same kind
        # of "this number cannot be right" evidence.
        if any(c.level == "sanity" and not c.passed for c in checks):
            verdict.dna_score = max(
                0.0, verdict.dna_score - float(cfg["dna"]["level1_range_violation"])
            )

        score_verdict(verdict, md, cfg)

    # ---------------- Stage 11b: derive what is still missing -------------- #
    confident = {
        k: v.candidate.value
        for k, v in verdicts.items()
        if v.candidate.value is not None and writes_to_excel(v.decision)
    }
    conf_map = {k: v.confidence for k, v in verdicts.items()}

    for md in metrics:
        if not md.applies_to(doc.inst_type):
            continue
        existing = verdicts.get(md.key)
        if existing and writes_to_excel(existing.decision):
            continue
        got = derivable(md.key, confident, conf_map, cfg)
        if not got:
            continue
        value, expression = got
        derived_cand = Candidate(
            metric=md.key,
            value=value,
            raw_text=f"derived: {expression}",
            page=0,
            section=(existing.candidate.section if existing else _best_section(pool, md.key)),
            section_score=1.0,
            year_label=target_fy,
            year_resolved=True,
            alias_matched="(derived)",
            alias_exact=True,
            from_table=False,
            derived=True,
            derivation=expression,
        )
        ctx = TrialContext(target_fy, prior_fy, md, doc.inst_type, cfg)
        dv = Verdict(
            candidate=derived_cand,
            defence_score=float(cfg["defence"]["base"])
            + float(cfg["defence"]["year_column_explicit"]),
            defence_reasons=[f"derived from confident inputs: {expression}"],
            court_score=float(cfg["defence"]["base"])
            + float(cfg["defence"]["year_column_explicit"]),
            source_score=1.0,
            dna_score=dna_score([level1_range(value, md)], cfg),
            formula_score=float(cfg["formulas"]["formula_pass_score"]),
            trend_score=trend_check(md.key, value, prior_values.get(md.key), cfg)[0],
            checks=[level1_range(value, md)],
        )
        # Derived values inherit the confidence of their weakest input, capped.
        inputs_conf = [
            conf_map.get(i, 0.0)
            for i in _formula_inputs(md.derivable_from)
        ]
        score_verdict(dv, md, cfg)
        dv.confidence = min(
            dv.confidence,
            min(inputs_conf) if inputs_conf else dv.confidence,
            float(cfg["formulas"]["derived_confidence_cap"]),
        )
        dv.decision = _decide(dv.confidence, cfg)
        dv.reason = f"derived ({expression}); " + dv.reason
        verdicts[md.key] = dv
        all_verdicts.setdefault(md.key, []).append(dv)

    # ---------------- Stage 16: build the cells ---------------------------- #
    for md in metrics:
        result.cells.append(
            _build_cell(
                md=md,
                doc=doc,
                result=result,
                verdict=verdicts.get(md.key),
                candidates=all_verdicts.get(md.key, []),
                pool_size=len(pool.get(md.key, [])),
                cfg=cfg,
            )
        )

    return result


FATAL_PROSECUTORS = ("year_mismatch", "unit_mismatch")


def _is_provisional(verdict: Verdict, md: MetricDef, write_floor: float) -> bool:
    """May this pass-1 winner be used as an input to the cross-metric checks?

    Two ways in:

    1. It already clears the write threshold on provenance alone. Obviously safe.
    2. It is *clean*: in range, and carrying no year-mismatch and no unit-mismatch
       penalty -- even if its confidence is low purely because it was found only
       once.

    Route 2 exists because of a genuine trap. A Tier-A metric that appears once
    (say CRAR, printed only in the Basel III section) scores badly on evidence.
    If we then gated it out of the provisional set, the ``CRAR = Tier 1 + Tier 2``
    identity -- the single best piece of corroboration available for it, and one
    that is completely independent of how many times it was printed -- would never
    run. The metric would be rejected *for lack of corroboration* while sitting
    next to the arithmetic that corroborates it.

    A candidate carrying a year or unit penalty is still excluded: those are the
    two ways a number can be plausibly *wrong*, and a wrong input poisons every
    formula it touches.
    """
    if verdict.candidate.value is None:
        return False
    if verdict.confidence >= write_floor:
        return True
    if not level1_range(verdict.candidate.value, md).passed:
        return False
    return not any(h.prosecutor in FATAL_PROSECUTORS for h in verdict.prosecutor_hits)


def _decide(confidence: float, cfg: dict) -> Decision:
    from arx.confidence import decide

    return decide(confidence, cfg)


def _formula_inputs(formula_id: Optional[str]) -> Sequence[str]:
    from arx.formulas import FORMULAS

    f = FORMULAS.get(formula_id or "")
    return f.inputs if f else ()


def _formula_metrics(formula_id: str) -> Sequence[str]:
    from arx.formulas import FORMULAS

    f = FORMULAS.get(formula_id)
    if not f:
        return ()
    return (f.target,) + tuple(f.inputs)


def _best_section(pool: Dict[str, List[Candidate]], key: str):
    from arx.models import Section

    cands = pool.get(key) or []
    return cands[0].section if cands else Section.UNKNOWN


def _prior_year_values(
    pool: Dict[str, List[Candidate]],
    prior_fy: Optional[str],
    cfg: dict,
) -> Dict[str, float]:
    """Best value per metric from the prior-year column -- for trend checks only."""
    if not prior_fy:
        return {}
    out: Dict[str, float] = {}
    for key, cands in pool.items():
        prior = [
            c
            for c in cands
            if c.year_label == prior_fy and c.year_resolved and c.value is not None
        ]
        if not prior:
            continue
        prior.sort(key=lambda c: (-c.section_score, -c.independent_sources, c.page))
        out[key] = prior[0].value
    return out


def _build_cell(
    md: MetricDef,
    doc: Document,
    result: DocumentResult,
    verdict: Optional[Verdict],
    candidates: Sequence[Verdict],
    pool_size: int,
    cfg: dict,
) -> CellResult:
    """One decided cell, with everything the audit trail needs."""
    nd = str(cfg["excel"]["nd_sentinel"])
    na = str(cfg["excel"]["na_sentinel"])

    if not md.applies_to(doc.inst_type):
        return CellResult(
            institution=result.institution,
            fiscal_year=result.fiscal_year,
            metric=md.key,
            excel_header=md.excel_header,
            sentinel=na,
            decision=Decision.NOT_APPLICABLE,
            confidence=100.0,
            reason=na_reason(md, doc.inst_type, result.category),
            candidate_count=pool_size,
        )

    if verdict is None or not writes_to_excel(verdict.decision):
        return CellResult(
            institution=result.institution,
            fiscal_year=result.fiscal_year,
            metric=md.key,
            excel_header=md.excel_header,
            sentinel=nd,
            decision=Decision.REJECT,
            confidence=verdict.confidence if verdict else 0.0,
            reason=rejection_reason(verdict, md, cfg),
            candidate_count=pool_size,
            pages=[verdict.candidate.page] if verdict else [],
            section=verdict.candidate.section.value if verdict else None,
            alias_matched=verdict.candidate.alias_matched if verdict else "",
            defence_score=verdict.defence_score if verdict else 0.0,
            prosecutor_hits=list(verdict.prosecutor_hits) if verdict else [],
            checks_passed=verdict.checks_passed if verdict else [],
            checks_failed=verdict.checks_failed if verdict else [],
            rejected=_rejected_list(candidates),
        )

    cand = verdict.candidate
    pages = sorted({cand.page} | {e.page for e in cand.evidence})
    return CellResult(
        institution=result.institution,
        fiscal_year=result.fiscal_year,
        metric=md.key,
        excel_header=md.excel_header,
        value=cand.value,
        text_value=cand.text_value,
        confidence=verdict.confidence,
        decision=verdict.decision,
        reason=verdict.reason,
        pages=[p for p in pages if p],
        section=cand.section.value,
        alias_matched=cand.alias_matched,
        unit_as_printed=cand.unit_as_printed,
        candidate_count=pool_size,
        defence_score=verdict.defence_score,
        prosecutor_hits=list(verdict.prosecutor_hits),
        checks_passed=verdict.checks_passed,
        checks_failed=verdict.checks_failed,
        formula_notes=[c.detail for c in verdict.checks if c.level == "formula"],
        derived=cand.derived,
        rejected=_rejected_list(candidates, winner=verdict),
    )


def _rejected_list(
    verdicts: Sequence[Verdict], winner: Optional[Verdict] = None, limit: int = 8
) -> List[str]:
    """``value@page`` for every candidate that did not win."""
    out: List[str] = []
    for v in verdicts:
        if winner is not None and v is winner:
            continue
        c = v.candidate
        if c.value is None and c.text_value is None:
            continue
        val = c.text_value if c.value is None else f"{c.value:g}"
        out.append(f"{val}@p{c.page}({v.court_score:.0f})")
    return out[:limit]


# --------------------------------------------------------------------------- #
# Batch driver
# --------------------------------------------------------------------------- #


def _process_one(args) -> dict:
    """ProcessPoolExecutor worker.

    Returns the :class:`DocumentResult` itself (pydantic models pickle cleanly,
    so there is no reason to round-trip through JSON and risk losing an enum on
    the way back).  A PDF that raises is caught here and reported -- one bad file
    must never kill the batch.
    """
    pdf_path, cfg, history, category_of, audit_dir, use_cache = args
    try:
        doc, pool = build_document(pdf_path, cfg, use_cache=use_cache)
        result = process_document(doc, pool, history, category_of, cfg)
        if audit_dir:
            dump_audit(result, audit_dir)
        return {"ok": True, "result": result, "path": str(pdf_path)}
    except Exception as exc:  # one bad PDF must never kill the batch
        return {
            "ok": False,
            "path": str(pdf_path),
            "error": f"{type(exc).__name__}: {exc}",
            "traceback": traceback.format_exc(),
        }


def dump_audit(result: DocumentResult, audit_dir: str | Path) -> Path:
    """Write ``audit/<institution>_<FY>.json`` with the FULL candidate pool.

    This is the file to open when you want to know why one specific cell came out
    the way it did: every candidate that was ever considered is in here.
    """
    audit_dir = Path(audit_dir)
    audit_dir.mkdir(parents=True, exist_ok=True)
    safe = "".join(ch if ch.isalnum() else "_" for ch in result.institution)
    path = audit_dir / f"{safe}_{result.fiscal_year}.json"
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(result.model_dump(mode="json"), fh, indent=2, default=str)
    return path


def run_batch(
    pdfs: Sequence[str | Path],
    template: str | Path,
    out_path: str | Path,
    workers: int = 1,
    cfg: Optional[dict] = None,
    use_cache: bool = True,
    progress: bool = True,
) -> tuple[Path, List[DocumentResult], List[dict]]:
    """The whole system, end to end.

    Returns ``(written_workbook, results, failures)``.  A PDF that blows up lands
    in ``failures`` with its traceback and the batch carries on.
    """
    cfg = cfg or load_config()
    seed_everything()

    audit_dir = Path(cfg["runtime"]["audit_dir"])
    history = read_history(template)
    category_of = _category_map(template)

    tasks = [
        (str(p), cfg, history, category_of, str(audit_dir), use_cache) for p in pdfs
    ]

    results: List[DocumentResult] = []
    failures: List[dict] = []

    if workers and workers > 1:
        with ProcessPoolExecutor(max_workers=workers) as pool_exec:
            futures = [pool_exec.submit(_process_one, t) for t in tasks]
            it = as_completed(futures)
            if progress:
                it = tqdm(it, total=len(futures), desc="PDFs", unit="pdf")
            for fut in it:
                _absorb(fut.result(), results, failures)
    else:
        it = tqdm(tasks, desc="PDFs", unit="pdf") if progress else tasks
        for task in it:
            _absorb(_process_one(task), results, failures)

    written = write_results(results, template, out_path, cfg)

    for fail in failures:
        log.error("FAILED %s: %s", fail["path"], fail["error"])
    log.info(
        "batch complete: %d succeeded, %d failed -> %s",
        len(results),
        len(failures),
        written,
    )
    return written, results, failures


def _absorb(payload: dict, results: List[DocumentResult], failures: List[dict]) -> None:
    if payload.get("ok"):
        results.append(payload["result"])
    else:
        failures.append(payload)


def _category_map(template: str | Path) -> Dict[str, str]:
    """``{institution: category}`` from the template's own column B."""
    from openpyxl import load_workbook

    cfg = load_config()
    xl = cfg["excel"]
    out: Dict[str, str] = {}
    path = Path(template)
    if not path.exists():
        return out
    wb = load_workbook(path, data_only=True, read_only=True)
    for name in wb.sheetnames:
        if not name.upper().startswith("FY"):
            continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=int(xl["first_data_row"]), values_only=True):
            if len(row) < 3:
                continue
            category, inst = row[1], row[2]
            if isinstance(inst, str) and inst.strip() and isinstance(category, str):
                out.setdefault(inst.strip(), category.strip())
    wb.close()
    return out


def discover_pdfs(folder: str | Path, recursive: bool = True) -> List[Path]:
    """Every PDF in a folder, sorted -- so a run is reproducible.

    Recursive discovery is enabled by default so users can point to a broad root
    (for example ``reports/`` with per-bank subfolders) without manual copying.
    """
    root = Path(folder)
    if not root.exists():
        return []

    walker = root.rglob("*.pdf") if recursive else root.glob("*.pdf")
    skip_dirs = {".git", ".venv", "venv", "node_modules", "__pycache__"}

    seen = set()
    out: List[Path] = []
    for p in walker:
        if any(part in skip_dirs for part in p.parts):
            continue
        key = str(p.resolve())
        if key in seen:
            continue
        seen.add(key)
        out.append(p)

    return sorted(out)
