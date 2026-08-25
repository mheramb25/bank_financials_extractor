"""Stage 16: the template-preserving Excel writer, and the audit sheets.

Rules this module enforces, in order of importance:

1. **We never rebuild the workbook.**  We copy the template file byte-for-byte
   and open the copy with openpyxl, so column widths, fonts, merges, freeze
   panes and the ``Source References`` sheet all survive untouched.
2. **Numbers are written as numbers.**  No currency symbol, no thousands
   separator, no percent sign, no string coercion. ``3.19`` means 3.19%.
3. **``ND`` never replaces a real number.**  A cell only changes if the new value
   is *more confident* than what is already there; both the old and the new are
   written into the Audit Trail when that happens.
4. **Every ``ND`` and every ``NA`` gets a row in Missing Cells Report** with a
   machine-generated reason.
5. Institutions are matched fuzzily on the name already in column C
   (``SBI`` == ``State Bank of India (SBI)``); an unknown institution is appended
   with the next Sr. No. and an inferred Category.
6. A fiscal year with no sheet gets one, cloned from the FY23-24 layout.
"""

from __future__ import annotations

import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz

from arx import load_config, load_institutions, load_metrics
from arx.models import CellResult, Decision, DocumentResult, MetricDef
from arx.normalize import normalize_name

log = logging.getLogger("arx.excel")

AUDIT_HEADERS = [
    "Institution",
    "FY",
    "Metric",
    "Chosen Value",
    "Unit as printed",
    "Page(s)",
    "Source Section",
    "Alias matched",
    "Candidate count",
    "Defence score",
    "Prosecutor penalties",
    "DNA checks passed/failed",
    "Formula checks",
    "Final confidence",
    "Decision (Auto/High/Manual/Reject)",
    "Reason for selection",
    "Rejected candidates (value@page, ...)",
]

SUMMARY_HEADERS = [
    "Institution",
    "Fiscal Year",
    "Cells filled",
    "Cells ND",
    "Cells NA",
    "Mean confidence",
    "Auto",
    "High",
    "Manual",
    "Reject",
    "Manual-Review cells",
]

MISSING_HEADERS = [
    "Institution",
    "Fiscal Year",
    "Metric",
    "Reason data unavailable",
]

GROUP_BANDS: List[Tuple[str, int, int]] = [
    ("Headline Size", 4, 11),
    ("Profit & Loss", 12, 23),
    ("Asset Quality", 24, 33),
    ("Capital & Solvency", 34, 43),
    ("Strategic & ESG", 44, 54),
    ("Technology", 55, 55),
]


# --------------------------------------------------------------------------- #
# Template construction (only used when you have no template, and by the tests)
# --------------------------------------------------------------------------- #


def build_blank_template(path: str | Path, years: Iterable[str] = ("FY21-22", "FY22-23", "FY23-24")) -> Path:
    """Create an empty workbook with exactly the template's layout.

    The real ``FinancialData_Verified_1.xlsx`` is the authority; this exists so
    the test-suite can run without shipping a copy of your workbook, and so a
    first-time user can bootstrap one.
    """
    metrics = load_metrics()
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=10)
    band_font = Font(bold=True, size=11, color="FFFFFF")
    band_fill = PatternFill("solid", fgColor="1F4E79")

    for year in years:
        ws = wb.create_sheet(year)
        ws.cell(row=1, column=1, value=f"Indian Banks & NBFCs — {year}  (₹ in Crore)")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        for name, c0, c1 in GROUP_BANDS:
            ws.cell(row=1, column=c0, value=name)
            if c1 > c0:
                ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        for cell in ws[1]:
            if cell.value:
                cell.font = band_font
                cell.fill = band_fill
                cell.alignment = Alignment(horizontal="center")

        for col, title in enumerate(["Sr. No.", "Category", "Name of Institution"], start=1):
            c = ws.cell(row=2, column=col, value=title)
            c.font = header_font
        for md in metrics:
            c = ws.cell(row=2, column=md.column, value=md.excel_header)
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="center")

        ws.freeze_panes = "D3"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 38
        for md in metrics:
            ws.column_dimensions[get_column_letter(md.column)].width = 18

    ref = wb.create_sheet("Source References")
    ref.cell(row=1, column=1, value="Institution")
    ref.cell(row=1, column=2, value="Fiscal Year")
    ref.cell(row=1, column=3, value="Source document")

    miss = wb.create_sheet("Missing Cells Report")
    for col, title in enumerate(MISSING_HEADERS, start=1):
        miss.cell(row=1, column=col, value=title).font = header_font

    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# The writer
# --------------------------------------------------------------------------- #


class TemplateWriter:
    """Writes :class:`DocumentResult` objects into a copy of the template."""

    def __init__(
        self,
        template: str | Path,
        out_path: str | Path,
        cfg: Optional[dict] = None,
    ) -> None:
        self.cfg = cfg or load_config()
        self.xl = self.cfg["excel"]
        self.template = Path(template)
        self.out_path = Path(out_path)

        if not self.template.exists():
            raise FileNotFoundError(f"template not found: {self.template}")

        self.out_path.parent.mkdir(parents=True, exist_ok=True)
        # Rule 1: copy the file, do not rebuild it.
        shutil.copyfile(self.template, self.out_path)
        self.wb = load_workbook(self.out_path)

        self.metrics: List[MetricDef] = load_metrics()
        self.by_key = {m.key: m for m in self.metrics}
        self.institutions = load_institutions()

        # (sheet, institution, metric) -> confidence written during THIS run.
        self._ledger: Dict[Tuple[str, str, str], float] = {}
        self._audit_rows: List[list] = []
        self._missing_rows: List[list] = []
        self._summary_rows: List[list] = []

    # -- sheet handling ---------------------------------------------------- #

    def _sheet_for(self, fiscal_year: str) -> Worksheet:
        """The sheet for this FY, cloning the layout donor if it does not exist."""
        if fiscal_year in self.wb.sheetnames:
            return self.wb[fiscal_year]

        donor_name = str(self.xl["clone_sheet_source"])
        donor = None
        if donor_name in self.wb.sheetnames:
            donor = self.wb[donor_name]
        else:
            for name in self.wb.sheetnames:
                if name.upper().startswith("FY"):
                    donor = self.wb[name]
                    break
        if donor is None:
            raise RuntimeError(
                "cannot create sheet %r: no FY sheet in the template to clone" % fiscal_year
            )

        log.info("cloning %s layout into new sheet %s", donor.title, fiscal_year)
        ws = self.wb.copy_worksheet(donor)
        ws.title = fiscal_year

        # Clear the donor's data rows but keep rows 1-2 (bands + headers).
        first = int(self.xl["first_data_row"])
        if ws.max_row >= first:
            ws.delete_rows(first, ws.max_row - first + 1)

        # Retitle the row-1 band so it names the right year.
        banner = ws.cell(row=1, column=1).value
        if isinstance(banner, str):
            ws.cell(
                row=1,
                column=1,
                value=re.sub(r"FY\d{2}-\d{2}", fiscal_year, banner),
            )
        return ws

    def _header_columns(self, ws: Worksheet) -> Dict[str, int]:
        """``{excel_header: column}`` read from row 2 of the sheet itself.

        The sheet is the authority, not ``metrics.yaml`` -- if you reorder the
        template's columns, we follow, and we never write into the wrong one.
        """
        out: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if isinstance(val, str) and val.strip():
                out[val.strip()] = col
        return out

    # -- institution rows -------------------------------------------------- #

    def _find_row(self, ws: Worksheet, institution: str) -> Optional[int]:
        """Fuzzy, alias-aware lookup of an institution's row in column C."""
        name_col = int(self.xl["name_column"])
        first = int(self.xl["first_data_row"])
        threshold = float(self.xl["fuzzy_name_match_threshold"])

        wanted = normalize_name(institution)
        aliases = {wanted}
        for inst in self.institutions:
            if normalize_name(inst.canonical) == wanted or institution in inst.aliases:
                aliases |= {normalize_name(a) for a in [inst.canonical] + inst.aliases}

        best_row, best_score = None, 0.0
        for row in range(first, ws.max_row + 1):
            cell = ws.cell(row=row, column=name_col).value
            if not isinstance(cell, str) or not cell.strip():
                continue
            existing = normalize_name(cell)
            for alias in aliases:
                if not alias:
                    continue
                if existing == alias:
                    return row
                score = float(fuzz.token_set_ratio(existing, alias))
                if score > best_score:
                    best_row, best_score = row, score

        return best_row if best_score >= threshold else None

    def _append_row(self, ws: Worksheet, institution: str, category: str) -> int:
        """Append a new institution with the next Sr. No. and inferred Category."""
        first = int(self.xl["first_data_row"])
        name_col = int(self.xl["name_column"])
        srno_col = int(self.xl["srno_column"])
        cat_col = int(self.xl["category_column"])

        row = first
        max_sr = 0
        for r in range(first, ws.max_row + 2):
            if not ws.cell(row=r, column=name_col).value:
                row = r
                break
            sr = ws.cell(row=r, column=srno_col).value
            if isinstance(sr, (int, float)):
                max_sr = max(max_sr, int(sr))
            row = r + 1

        ws.cell(row=row, column=srno_col, value=max_sr + 1)
        ws.cell(row=row, column=cat_col, value=category)
        ws.cell(row=row, column=name_col, value=institution)
        log.info("appended %s to %s at row %d (Sr. No. %d)", institution, ws.title, row, max_sr + 1)
        return row

    def _category_for(self, result: DocumentResult) -> str:
        """Category from the alias table, else the default for the type."""
        for inst in self.institutions:
            if inst.canonical == result.institution:
                return inst.category
        defaults = {
            "bank": "Private Sector Bank",
            "sfb": "Small Finance Bank",
            "nbfc": "NBFC",
            "hfc": "Housing Finance Company",
            "aifi": "All-India Financial Institution",
        }
        return result.category or defaults.get(result.inst_type, "NBFC")

    # -- the actual write -------------------------------------------------- #

    def write_result(self, result: DocumentResult) -> None:
        """Write one PDF's worth of cells, plus its audit and missing-cell rows."""
        ws = self._sheet_for(result.fiscal_year)
        headers = self._header_columns(ws)
        row = self._find_row(ws, result.institution)
        if row is None:
            row = self._append_row(ws, result.institution, self._category_for(result))
        else:
            # Keep the template's own spelling of the name; just refresh Category
            # if the template left it blank.
            cat_col = int(self.xl["category_column"])
            if not ws.cell(row=row, column=cat_col).value:
                ws.cell(row=row, column=cat_col, value=self._category_for(result))

        nd = str(self.xl["nd_sentinel"])
        na = str(self.xl["na_sentinel"])
        comment_below = float(self.xl["add_cell_comment_below"])
        assumed = float(self.xl["assumed_existing_confidence"])

        for cell_res in result.cells:
            col = headers.get(cell_res.excel_header)
            if col is None:
                log.warning(
                    "header %r not found in sheet %s -- skipping",
                    cell_res.excel_header,
                    ws.title,
                )
                continue

            target = ws.cell(row=row, column=col)
            existing = target.value
            ledger_key = (ws.title, result.institution, cell_res.metric)
            existing_conf = self._ledger.get(ledger_key, assumed)
            existing_is_real = existing not in (None, "", nd, na)

            new_value = cell_res.excel_value
            new_is_real = cell_res.is_written_number

            # Rule 3: never downgrade a real number to a sentinel.
            if not new_is_real and existing_is_real:
                self._audit(
                    result,
                    cell_res,
                    note=(
                        f"kept existing value {existing!r}: refused to overwrite a real "
                        f"number with {new_value!r}"
                    ),
                )
                self._missing(result, cell_res, kept_existing=True)
                continue

            # Rule 3, second half: only upgrade on strictly higher confidence.
            if (
                new_is_real
                and existing_is_real
                and bool(self.xl["overwrite_only_if_more_confident"])
                and cell_res.confidence <= existing_conf
            ):
                self._audit(
                    result,
                    cell_res,
                    note=(
                        f"kept existing value {existing!r} (confidence {existing_conf:.0f}) "
                        f"over new {new_value!r} (confidence {cell_res.confidence:.0f})"
                    ),
                )
                continue

            target.value = new_value
            if new_is_real:
                self._ledger[ledger_key] = cell_res.confidence
                if cell_res.confidence < comment_below:
                    target.comment = Comment(
                        f"arx confidence {cell_res.confidence:.0f} "
                        f"({cell_res.decision.value})\n{cell_res.reason}",
                        "arx",
                    )
            self._audit(
                result,
                cell_res,
                note=(
                    f"overwrote {existing!r} (assumed confidence {existing_conf:.0f})"
                    if existing_is_real and new_is_real
                    else ""
                ),
            )
            if cell_res.sentinel:
                self._missing(result, cell_res)

        self._summarise(result)

    # -- audit sheets ------------------------------------------------------ #

    def _audit(self, result: DocumentResult, cell: CellResult, note: str = "") -> None:
        penalties = "; ".join(
            f"{h.prosecutor} -{h.penalty:g}: {h.reason}" for h in cell.prosecutor_hits
        )
        checks = (
            f"passed: {', '.join(cell.checks_passed) or '-'} | "
            f"failed: {', '.join(cell.checks_failed) or '-'}"
        )
        self._audit_rows.append(
            [
                result.institution,
                result.fiscal_year,
                cell.excel_header,
                cell.excel_value,
                cell.unit_as_printed,
                ", ".join(str(p) for p in cell.pages),
                cell.section or "",
                cell.alias_matched,
                cell.candidate_count,
                round(cell.defence_score, 1),
                penalties,
                checks,
                "; ".join(cell.formula_notes),
                round(cell.confidence, 1),
                cell.decision.value,
                (cell.reason + (f" || {note}" if note else "")),
                ", ".join(cell.rejected),
            ]
        )

    def _missing(
        self, result: DocumentResult, cell: CellResult, kept_existing: bool = False
    ) -> None:
        reason = cell.reason
        if kept_existing:
            reason = f"{reason} (existing template value retained)"
        self._missing_rows.append(
            [result.institution, result.fiscal_year, cell.excel_header, reason]
        )

    def _summarise(self, result: DocumentResult) -> None:
        filled = [c for c in result.cells if c.is_written_number]
        nd = [c for c in result.cells if c.sentinel == self.xl["nd_sentinel"]]
        na = [c for c in result.cells if c.sentinel == self.xl["na_sentinel"]]
        manual = [c for c in filled if c.decision == Decision.MANUAL]
        bands = {d.value: 0 for d in Decision}
        for c in result.cells:
            bands[c.decision.value] = bands.get(c.decision.value, 0) + 1

        mean_conf = (
            sum(c.confidence for c in filled) / len(filled) if filled else 0.0
        )
        self._summary_rows.append(
            [
                result.institution,
                result.fiscal_year,
                len(filled),
                len(nd),
                len(na),
                round(mean_conf, 1),
                bands.get("Auto", 0),
                bands.get("High", 0),
                bands.get("Manual", 0),
                bands.get("Reject", 0),
                ", ".join(c.excel_header for c in manual),
            ]
        )

    def _write_sheet(
        self, name: str, headers: List[str], rows: List[list], append: bool = False
    ) -> None:
        """Create (or append to) one of the audit sheets."""
        if name in self.wb.sheetnames:
            ws = self.wb[name]
            if not append:
                first = 2
                if ws.max_row >= first:
                    ws.delete_rows(first, ws.max_row - first + 1)
        else:
            ws = self.wb.create_sheet(name)
        if ws.max_row == 1 and not ws.cell(row=1, column=1).value:
            for col, title in enumerate(headers, start=1):
                c = ws.cell(row=1, column=col, value=title)
                c.font = Font(bold=True)
            ws.freeze_panes = "A2"
        start = ws.max_row + 1 if ws.cell(row=1, column=1).value else 2
        for r, row in enumerate(rows, start=start):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def save(self) -> Path:
        """Flush the audit sheets and write the workbook."""
        self._write_sheet(
            str(self.xl["missing_sheet"]), MISSING_HEADERS, self._missing_rows, append=True
        )
        self._write_sheet(str(self.xl["audit_sheet"]), AUDIT_HEADERS, self._audit_rows)
        self._write_sheet(
            str(self.xl["summary_sheet"]), SUMMARY_HEADERS, self._summary_rows
        )

        meta = self.wb[str(self.xl["summary_sheet"])]
        meta.cell(
            row=meta.max_row + 2,
            column=1,
            value=f"Generated by arx (offline) at {datetime.now():%Y-%m-%d %H:%M:%S}",
        )

        self.wb.save(self.out_path)
        log.info(
            "wrote %s (%d audit rows, %d missing-cell rows)",
            self.out_path,
            len(self._audit_rows),
            len(self._missing_rows),
        )
        return self.out_path


def write_results(
    results: Iterable[DocumentResult],
    template: str | Path,
    out_path: str | Path,
    cfg: Optional[dict] = None,
) -> Path:
    """Convenience: write a whole batch and save.

    Results are sorted by (fiscal year, institution) so that the same batch always
    produces byte-comparable output regardless of the order the PDFs finished in.
    """
    writer = TemplateWriter(template, out_path, cfg)
    for result in sorted(results, key=lambda r: (r.fiscal_year, r.institution)):
        writer.write_result(result)
    return writer.save()


def read_history(
    workbook: str | Path,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """``{institution: {metric_key: {fy: value}}}`` from an existing workbook.

    This is the Level-3 (historical) and Level-4 (peer) DNA input: the years we
    have already established become the yardstick for the year we are extracting.
    """
    cfg = load_config()
    xl = cfg["excel"]
    metrics = load_metrics()
    header_to_key = {m.excel_header: m.key for m in metrics}

    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    path = Path(workbook)
    if not path.exists():
        return out

    wb = load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        if not name.upper().startswith("FY"):
            continue
        ws = wb[name]
        cols: Dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col).value
            if isinstance(header, str) and header.strip() in header_to_key:
                cols[col] = header_to_key[header.strip()]

        for row in range(int(xl["first_data_row"]), ws.max_row + 1):
            inst = ws.cell(row=row, column=int(xl["name_column"])).value
            if not isinstance(inst, str) or not inst.strip():
                continue
            bucket = out.setdefault(inst.strip(), {})
            for col, key in cols.items():
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    bucket.setdefault(key, {})[name] = float(val)
    wb.close()
    return out


def peer_values(
    history: Dict[str, Dict[str, Dict[str, float]]],
    category_of: Dict[str, str],
    category: str,
    fiscal_year: str,
    exclude: str,
) -> Dict[str, List[float]]:
    """``{metric: [peer values]}`` for peers of the same category in the same FY."""
    out: Dict[str, List[float]] = {}
    for inst, metrics in history.items():
        if inst == exclude or category_of.get(inst) != category:
            continue
        for key, by_year in metrics.items():
            val = by_year.get(fiscal_year)
            if val is not None:
                out.setdefault(key, []).append(val)
    return out
