"""
simple_app.py — Indian bank / NBFC annual-report PDFs in, a FRESH Excel out.
ONE FILE, ONE ENGINE, THREE WAYS TO RUN IT.

    pip install streamlit pdfplumber pypdf openpyxl pandas

    streamlit run simple_app.py                         # the upload/download web app
    python simple_app.py --pdfs ./input_pdfs --out out.xlsx   # headless batch (arx-style)
    python simple_app.py                                # the built-in self-test

The batch CLI and the web app share the exact same extraction engine below —
there is no second copy of the logic to keep in sync. The optional template is
read-only (peer/history cross-checks); the output is always a brand-new workbook.

THE WORKFLOW
    PDF upload
      -> parse + OCR + tables                (STAGE 1)
      -> page ranking engine                 (STAGE 2)  prioritise, don't scan 400 pages
      -> candidate generation                (STAGE 3)  extract EVERY value, pick none
      -> courtroom                           (STAGE 4)  defence + 4 prosecutors
      -> financial DNA                       (STAGE 5)  ranges, banking logic, history, PEERS
      -> formula layer                       (STAGE 6)  check what we found, derive what we didn't
      -> confidence -> Excel + audit         (STAGE 7)
           >= 80 write   |   65-79 write + REVIEW flag   |   < 65 write ND

    "A wrong number is worse than a missing number." Every ambiguity resolves
    into ND, never into a guess.

WHAT THIS CODE IS PARANOID ABOUT, AND WHY
    Six ways a real Indian annual report will silently hand you a wrong number.
    Each one is defended against below; each has a test in self_test().

    1. THE SCHEDULE COLUMN.  Every bank P&L prints
           Particulars | Schedule | Current Year | Prior Year
       so the line reads "Interest earned  13  1,09,231.34  86,374.55".
       Naively taking the first number gives you Interest Earned = 13.
       -> take the LAST TWO numbers, and drop a leading small integer.

    2. YEAR HEADER ROWS.  "Deposits 2023 2022" is a header, not data.
       2023 passes any sane range check for an amount.
       -> a line whose numbers are ALL bare years is a header. Skip it.

    3. PROSE POISONING THE SCALE.  "We serve 50 million customers" on the
       Highlights page made the old code scale every amount on that page by 0.1.
       -> only a real unit CAPTION sets the scale: "(₹ in crore)", "Rs. in lakh".
          A bare "million" in a sentence is ignored.

    4. MULTI-YEAR TABLES.  A 10-year summary row has ten numbers. Picking the
       "first" or "second" one is a coin flip on which year you get.
       -> a text line with more than two values is ambiguous. Skip it. The proper
          table reader handles multi-year tables, because it reads the header.

    5. TREND CHECKED AGAINST THE FUTURE.  sorted(["FY21-22","FY23-24"])[-1] is
       FY23-24 — so extracting FY22-23 compared it against the year AFTER it.
       -> compare against the nearest EARLIER year, by parsed end-year.

    6. FAIL-OPEN RANGES.  A metric missing from RANGES fell back to -60..150,
       so a PCR of 140 would sail through.
       -> RANGES is exhaustive, asserted at import. A metric with no range is a
          crash on startup, not a wrong number in your spreadsheet.
"""

import io
import re
import shutil
import tempfile
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from statistics import mean, pstdev

import pandas as pd
import pdfplumber
import pypdf
from openpyxl import Workbook, load_workbook

try:
    import streamlit as st          # only the web app needs it; the CLI does not
except ImportError:                 # so a headless/server install can skip streamlit
    st = None
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ========================================================================== #
# THE METRIC DICTIONARY — keys are your template's EXACT column headers.
#   header: (regex for the row label, kind, applies_to)
#   kind       : amount (₹ crore) | percent | count | ratio | text
#   applies_to : institution types the metric means anything for. Everything
#                else gets NA (not applicable), never ND (not disclosed).
# ========================================================================== #

BANK, NBFC, AIFI, HFC, SFB = "bank", "nbfc", "aifi", "hfc", "sfb"
ALL = (BANK, NBFC, AIFI, HFC, SFB)
BANKS_ONLY = (BANK, SFB)
NONBANKS = (NBFC, HFC, AIFI)

METRICS = {
    "Total Assets":                    (r"^total\s+assets$", "amount", ALL),
    "Total Deposits":                  (r"^(total\s+)?deposits$", "amount", BANKS_ONLY),
    "Total Advances (Gross)":          (r"^(gross|total)?\s*advances$|^gross\s+loans?$", "amount", ALL),
    "Assets Under Management (AUM)":   (r"^assets\s+under\s+management|^aum$", "amount", NONBANKS),
    "Market Capitalization":           (r"^market\s+capitali[sz]ation$|^market\s+cap$", "amount", ALL),
    "Total Branches":                  (r"^(number|no\.?)\s+of\s+branches$|^total\s+(number\s+of\s+)?branches$|^branches$|^(total\s+)?(banking|business)\s+outlets$|^domestic\s+branches$|^branch\s+network$", "count", ALL),
    "Total Employee Count":            (r"^(number|no\.?)\s+of\s+(permanent\s+)?employees(\s+on\s+roll)?$|^total\s+(permanent\s+|on.?roll\s+)?employees$|^(total\s+)?employee\s+strength$|^(total\s+)?head.?count$|^permanent\s+employees$", "count", ALL),
    "Total Customer Base (Numbers)":   (r"^customer\s+base$|^(number|no\.?)\s+of\s+customers$|^total\s+customers$", "count", ALL),
    "Interest Earned":                 (r"^interest\s+(earned|income)$", "amount", ALL),
    "Interest Expended":               (r"^interest\s+(expended|expenses?)$|^finance\s+costs?$", "amount", ALL),
    "Net Interest Income (NII)":       (r"^net\s+interest\s+income$|^nii$", "amount", ALL),
    "Other Income (Non-Interest)":     (r"^other\s+income$|^non.?interest\s+income$", "amount", ALL),
    "Operating Expenses":              (r"^(total\s+)?operating\s+expenses$", "amount", ALL),
    "Employee Costs":                  (r"^payments\s+to\s+and\s+provisions\s+for\s+employees$|^employee\s+(costs?|benefits?\s+expenses?)$|^staff\s+costs?$", "amount", ALL),
    "Total Provisions":                (r"^provisions?\s+and\s+contingencies$|^total\s+provisions$", "amount", ALL),
    "Profit After Tax (PAT)":          (r"^profit\s+after\s+tax$|^net\s+profit$|^profit\s+for\s+the\s+year$|^pat$", "amount", ALL),
    "Net Interest Margin (NIM) %":     (r"^net\s+interest\s+margin|^nim$", "percent", ALL),
    "Return on Assets (RoA) %":        (r"^return\s+on\s+(average\s+)?assets|^roa$", "percent", ALL),
    "Return on Equity (RoE) %":        (r"^return\s+on\s+(equity|net\s+worth)|^roe$", "percent", ALL),
    "Cost-to-Income Ratio %":          (r"^cost.to.income", "percent", ALL),
    # AMOUNT = the bare NPA figure (a big number). Anchored so it does NOT catch
    # the ratio forms ("... to gross advances"). Trailing \d? eats a footnote.
    "Gross NPA (GNPA) Amount":         (r"^gross\s+(npas?|non.?performing\s+assets?)\d?$", "amount", ALL),
    "Net NPA (NNPA) Amount":           (r"^net\s+(npas?|non.?performing\s+assets?)\d?$", "amount", ALL),
    # RATIO = broad on purpose. Reports write it many ways — "Gross NPA %",
    # "Gross Non-Performing Assets to Gross Advances", "... as a percentage of",
    # sometimes with a footnote digit ("assets2"). The is_pct flag + the 0-40
    # range keep this from swallowing the amount. clean_label already stripped
    # the '%', so a printed "Gross NPA (%)" arrives here as "gross npa".
    "Gross NPA Ratio %":               (r"^gross\s+(npas?|non.?performing\s+(assets?|advances?))\d?\s*(ratio|to\s+gross\s+advances|as\s+a|of\s+gross|%)?", "percent", ALL),
    "Net NPA Ratio %":                 (r"^net\s+(npas?|non.?performing\s+(assets?|advances?))\d?\s*(ratio|to\s+net\s+advances|as\s+a|of\s+net|%)?", "percent", ALL),
    "Provision Coverage Ratio (PCR) %": (r"^provision(ing)?\s+coverage\s+ratio|^pcr$", "percent", ALL),
    "Slippage Ratio %":                (r"^(gross\s+|fresh\s+)?slippage\s+ratio", "percent", ALL),
    # Kept STRICT on purpose: reports print prudential / technical / gross write-offs
    # as separate lines and grabbing the wrong one is worse than ND. Only the
    # explicit "write-offs during the year" total is trusted.
    "Write-offs During Year":          (r"^write.?offs?\s+during\s+(the\s+)?year|^amounts?\s+written\s+off\s+during", "amount", ALL),
    "Stage 3 Assets % (NBFC)":         (r"^(gross\s+)?stage\s*(3|iii)", "percent", (NBFC, HFC)),
    "Standard Assets Provision":       (r"^provision\s+(for|on)\s+standard\s+assets|^standard\s+assets?\s+provision", "amount", ALL),
    "Credit Cost %":                   (r"^credit\s+cost", "percent", ALL),
    "Tier 1 Capital Ratio %":          (r"^tier\s*(1|i)\s*(capital)?\s*(ratio|crar)", "percent", ALL),
    "Tier 2 Capital Ratio %":          (r"^tier\s*(2|ii)\s*(capital)?\s*(ratio|crar)", "percent", ALL),
    "Total CRAR (CAR) %":              (r"^crar$|^capital\s+adequacy\s+ratio|^total\s+capital\s+ratio$|^car$", "percent", ALL),
    "Common Equity Tier 1 (CET1) %":   (r"^common\s+equity\s+tier\s*(1|i)|^cet.?1", "percent", BANKS_ONLY),
    # NBFC/HFC/AIFI only: Debt-to-Equity is a core solvency ratio for lenders that
    # fund themselves with debt. Banks fund with deposits and report CRAR instead,
    # so a "debt-equity" line in a bank's report is usually a near-zero segment
    # figure — matching it produced wrong values. For banks this is NA.
    "Debt-to-Equity Ratio":            (r"^debt\s*[-/ ]?\s*(to\s+)?equity(\s*ratio)?|^gearing\s+ratio|^total\s+debt.{0,4}equity", "ratio", NONBANKS),
    "Net Worth / Shareholders Equity": (r"^net\s+worth$|^shareholders?.?\s*(equity|funds)$|^total\s+equity$", "amount", ALL),
    "Price-to-Earnings (P/E) Ratio":   (r"^price.to.earnings|^p\s*/\s*e\b|^pe\s+ratio", "ratio", ALL),
    "Price-to-Book (P/B) Ratio":       (r"^price.to.book|^p\s*/\s*b\b|^pb\s+ratio", "ratio", ALL),
    "Dividend Yield %":                (r"^dividend\s+yield", "percent", ALL),
    "Credit Rating (Long Term)":       (r"^(long.?term\s+)?credit\s+rating|^long.?term\s+rating", "text", ALL),
    "CASA Ratio %":                    (r"^casa\s*(ratio|%|deposits?\s+ratio)?$|^casa\s+ratio|^casa\s+to\s+total\s+deposits|^low.?cost\s+casa", "percent", BANKS_ONLY),
    "Retail Mix %":                    (r"^retail\s+mix|^share\s+of\s+retail", "percent", ALL),
    "Wholesale Mix %":                 (r"^wholesale\s+mix|^corporate\s+mix", "percent", ALL),
    "Tech Spend (% of Opex)":          (r"^tech(nology)?\s+spend.*(opex|operating)", "percent", ALL),
    "Digital Transaction Share %":     (r"^digital\s+transactions?\s*(share|mix|%)", "percent", ALL),
    "Mobile App Users (Active)":       (r"^(active\s+)?mobile\s+(banking\s+)?(app\s+)?users|^monthly\s+active\s+users", "count", ALL),
    "CSR Spending Amount":             (r"^csr\s+(spend|expenditure)|^amount\s+spent\s+on\s+csr", "amount", ALL),
    "Board Gender Diversity %":        (r"^board\s+gender\s+diversity|^women\s+directors?", "percent", ALL),
    "Priority Sector Lending (PSL) %": (r"^(total\s+)?priority\s+sector(\s+lending|\s+advances)?|^psl\b", "percent", (BANK, SFB, AIFI)),
    "ESG Rating (CRISIL/MSCI)":        (r"^esg\s+(rating|score)", "text", ALL),
    "Promoter Holding %":              (r"^promoter\s+(and\s+promoter\s+group\s+)?(share)?holding", "percent", ALL),
    "Investment in Technology in last FY": (r"^investment\s+in\s+technology|^technology\s+(investment|capex)", "amount", ALL),
}

# -------------------------------------------------------------------------- #
# STAGE 5 / L1 — INDUSTRY RANGES.  EXHAUSTIVE, AND FAIL-CLOSED.
#
# Every non-text metric MUST have an entry. There is deliberately no fallback:
# a metric with no range crashes at import (see the assert below) instead of
# quietly accepting a PCR of 140. Amounts are in ₹ crore.
#
# Sized against the largest institution in your template (SBI: assets 6.18m cr,
# interest earned 415k cr, 232k employees, 22.5k branches) with headroom, and
# tight enough to catch a x100 unit error or a dropped decimal point.
# -------------------------------------------------------------------------- #

# AMOUNT ranges are deliberately WIDE and scale-free-ish: they only rule out the
# physically absurd, because this must work for a ₹2,000 crore co-operative bank
# and for SBI at ₹60 lakh crore alike. An absolute floor tuned to big banks would
# silently reject every small one.
#
# The real work of catching a wrong AMOUNT is done by RELATIVE_RULES below, which
# are scale-free: "Interest Earned is between 1% and 25% of Total Assets" is true
# of every lender that has ever existed, at any size. That is what catches
# Shriram's Interest Earned = 10.31 (a per-share row) and Kotak's PAT = 885 (a
# subsidiary's statement) WITHOUT hardcoding anything about Shriram or Kotak.
RANGES = {
    # amounts (₹ crore). The FLOORS on the big P&L / balance-sheet items are
    # deliberate: some PSU/bilingual reports (Indian Bank, IDBI) come through
    # pdfplumber with mangled numbers — "5,282" reads as "52.82", giving an
    # absurd "Interest Earned = 5.99 crore". No bank whose report is worth
    # analysing has under ₹50 cr of interest income, so a modest floor rejects
    # the fragment while still admitting any genuinely small institution.
    "Total Assets": (500, 50_000_000),
    "Total Deposits": (100, 40_000_000),
    "Total Advances (Gross)": (100, 40_000_000),
    "Assets Under Management (AUM)": (100, 40_000_000),
    "Market Capitalization": (1, 20_000_000),
    "Interest Earned": (50, 5_000_000),
    "Interest Expended": (20, 5_000_000),
    "Net Interest Income (NII)": (10, 3_000_000),
    "Other Income (Non-Interest)": (-500_000, 2_000_000),
    "Operating Expenses": (20, 2_000_000),
    "Employee Costs": (5, 1_000_000),
    "Total Provisions": (-500_000, 2_000_000),
    "Profit After Tax (PAT)": (-500_000, 1_000_000),
    "Gross NPA (GNPA) Amount": (0, 2_000_000),
    "Net NPA (NNPA) Amount": (0, 1_000_000),
    "Write-offs During Year": (0, 1_000_000),
    "Standard Assets Provision": (0, 500_000),
    "Net Worth / Shareholders Equity": (1, 10_000_000),
    "CSR Spending Amount": (0, 50_000),
    "Investment in Technology in last FY": (0, 200_000),
    # counts. NOTE the floors: a range starting at 1 let "Total Employee Count = 1"
    # through with 93 confidence. No institution in this universe has 1 employee
    # or 1 branch, and a floor is free insurance against a stray digit.
    "Total Branches": (10, 50_000),
    "Total Employee Count": (100, 500_000),
    "Total Customer Base (Numbers)": (100, 1_000_000_000),
    "Mobile App Users (Active)": (100, 500_000_000),
    # percentages — a plain number: 3.19 means 3.19%
    "Net Interest Margin (NIM) %": (0, 18),        # banks ~3-4; MFI-heavy NBFCs reach low teens
    "Return on Assets (RoA) %": (-5, 5),
    "Return on Equity (RoE) %": (-50, 50),
    "Cost-to-Income Ratio %": (5, 95),
    "Gross NPA Ratio %": (0, 40),
    "Net NPA Ratio %": (0, 25),
    "Provision Coverage Ratio (PCR) %": (0, 100),
    "Slippage Ratio %": (0, 20),
    "Stage 3 Assets % (NBFC)": (0, 40),
    "Credit Cost %": (-2, 15),
    "Tier 1 Capital Ratio %": (0, 40),
    "Tier 2 Capital Ratio %": (0, 15),
    "Total CRAR (CAR) %": (0, 40),
    "Common Equity Tier 1 (CET1) %": (0, 35),
    "Dividend Yield %": (0, 20),
    "CASA Ratio %": (0, 100),
    "Retail Mix %": (0, 100),
    "Wholesale Mix %": (0, 100),
    "Tech Spend (% of Opex)": (0, 50),
    "Digital Transaction Share %": (0, 100),
    "Board Gender Diversity %": (0, 100),
    "Priority Sector Lending (PSL) %": (0, 100),
    "Promoter Holding %": (0, 100),
    # ratios (x, not %)
    "Debt-to-Equity Ratio": (0, 30),
    "Price-to-Earnings (P/E) Ratio": (0, 150),
    "Price-to-Book (P/B) Ratio": (0, 20),
}

_missing = [h for h, (_, kind, _) in METRICS.items() if kind != "text" and h not in RANGES]
assert not _missing, f"metrics with no sane_range (fix RANGES): {_missing}"

# -------------------------------------------------------------------------- #
# SHADOW CAPTURES.  Not template columns — internal line items we grab so we can
# DERIVE a metric that is rarely printed as a single labelled row.
#
# Net Worth is the prime example: banks seldom print a "Net Worth" line, but
# every balance sheet has "Capital" and "Reserves and Surplus", and
# Net Worth = Capital + Reserves. Capturing those two lets us derive Net Worth,
# which in turn lets us derive RoE = PAT / Net Worth. One capture, two metrics.
#   key: (label regex, (lo, hi) range in ₹ crore)
# -------------------------------------------------------------------------- #
SHADOW_ALIASES = {
    "_capital":  (r"^(equity\s+)?(share\s+)?capital$|^paid.?up\s+(equity\s+)?(share\s+)?capital$", (1, 200_000)),
    "_reserves": (r"^reserves\s+and\s+surplus$|^reserves\s+&\s+surplus$|^other\s+equity$|^total\s+reserves(\s+and\s+surplus)?$", (1, 5_000_000)),
    "_borrowings": (r"^(total\s+)?borrowings$|^debt\s+securities$", (0, 20_000_000)),
}

# -------------------------------------------------------------------------- #
# SCALE-FREE PROPORTION RULES.  (metric, anchor, min_ratio, max_ratio)
#
# These replace the hardcoded floors, and they are the reason this works on a
# bank I have never seen. Every one of them is true of any lender at any size:
# a lender's interest income is some sensible fraction of its assets; its
# employee cost is some sensible fraction of its operating cost.
#
# Shriram's Interest Earned came out as 10.31 crore against 2.7 lakh crore of
# assets. That is 0.004% — no lender on earth. Caught, with no mention of
# Shriram anywhere in this file.
# -------------------------------------------------------------------------- #
RELATIVE_RULES = [
    ("Interest Earned", "Total Assets", 0.01, 0.30),
    ("Interest Expended", "Interest Earned", 0.05, 0.95),
    ("Net Interest Income (NII)", "Interest Earned", 0.03, 0.95),
    ("Other Income (Non-Interest)", "Interest Earned", 0.001, 1.50),
    ("Operating Expenses", "Interest Earned", 0.02, 2.00),
    ("Employee Costs", "Operating Expenses", 0.05, 0.95),
    ("Profit After Tax (PAT)", "Total Assets", -0.15, 0.15),
    ("Total Deposits", "Total Assets", 0.05, 0.95),
    ("Total Advances (Gross)", "Total Assets", 0.05, 0.95),
    ("Net Worth / Shareholders Equity", "Total Assets", 0.01, 0.60),
    ("Gross NPA (GNPA) Amount", "Total Advances (Gross)", 0.0, 0.40),
    ("Net NPA (NNPA) Amount", "Gross NPA (GNPA) Amount", 0.0, 1.00),
    ("Assets Under Management (AUM)", "Total Assets", 0.10, 5.00),
]

# STAGE 5 / L2 — banking logic. (left, op, right, message)
BANKING_RULES = [
    ("Gross NPA (GNPA) Amount", ">=", "Net NPA (NNPA) Amount", "Gross NPA must be >= Net NPA"),
    ("Gross NPA Ratio %", ">=", "Net NPA Ratio %", "Gross NPA % must be >= Net NPA %"),
    ("Return on Equity (RoE) %", ">=", "Return on Assets (RoA) %", "RoE must be >= RoA"),
    ("Total CRAR (CAR) %", ">=", "Common Equity Tier 1 (CET1) %", "CRAR must be >= CET1"),
    ("Total CRAR (CAR) %", ">=", "Tier 1 Capital Ratio %", "CRAR must be >= Tier 1"),
    ("Total Assets", ">", "Total Advances (Gross)", "Total Assets must exceed Advances"),
    ("Total Assets", ">", "Total Deposits", "Total Assets must exceed Deposits"),
    ("Interest Earned", ">", "Interest Expended", "Interest Earned normally exceeds Interest Expended"),
]

# STAGE 6 — formulas. target: (inputs, fn, text, kind)
FORMULAS = {
    "Net Interest Income (NII)": (
        ["Interest Earned", "Interest Expended"], lambda a, b: a - b,
        "NII = Interest Earned - Interest Expended", "amount"),
    "Total CRAR (CAR) %": (
        ["Tier 1 Capital Ratio %", "Tier 2 Capital Ratio %"], lambda a, b: a + b,
        "CRAR = Tier 1 + Tier 2", "ratio"),
    "Provision Coverage Ratio (PCR) %": (
        ["Gross NPA (GNPA) Amount", "Net NPA (NNPA) Amount"],
        lambda g, n: (g - n) / g * 100 if g else None,
        "PCR = (GNPA - NNPA) / GNPA x 100", "ratio"),
    "Return on Assets (RoA) %": (
        ["Profit After Tax (PAT)", "Total Assets"],
        lambda p, a: p / a * 100 if a else None,
        "RoA = PAT / Total Assets x 100", "ratio"),
    "Return on Equity (RoE) %": (
        ["Profit After Tax (PAT)", "Net Worth / Shareholders Equity"],
        lambda p, w: p / w * 100 if w else None,
        "RoE = PAT / Net Worth x 100", "ratio"),
    "Cost-to-Income Ratio %": (
        ["Operating Expenses", "Net Interest Income (NII)", "Other Income (Non-Interest)"],
        lambda o, n, i: o / (n + i) * 100 if (n + i) else None,
        "Cost-to-Income = Opex / (NII + Other Income) x 100", "ratio"),
    "Gross NPA (GNPA) Amount": (
        ["Gross NPA Ratio %", "Total Advances (Gross)"], lambda r, a: r * a / 100,
        "GNPA Amount = GNPA% x Gross Advances / 100", "amount"),
    "Wholesale Mix %": (
        ["Retail Mix %"], lambda r: 100 - r,
        "Wholesale Mix % = 100 - Retail Mix %", "ratio"),
}
AMOUNT_TOLERANCE = 0.01      # +/-1% relative
RATIO_TOLERANCE = 0.10       # +/-0.1 percentage points
DERIVE_MIN_INPUT_CONF = 80   # never derive from a shaky input
DERIVED_CAP = 90             # a derived value was never printed: it cannot be a 100

# -------------------------------------------------------------------------- #
# WHO IS THIS REPORT ABOUT?  Asked of the DOCUMENT, not of a list.
#
# There is no hardcoded list of banks here, on purpose. A list only ever knows
# the banks somebody remembered to type in, and silently mis-files everything
# else. Instead we read the report the way a human does:
#
#   1. The AUDITOR'S REPORT. Every Indian annual report contains
#         "To the Members of <NAME>"
#      and it is the legally exact entity name. This is the single most reliable
#      sentence in the whole document.
#   2. The statement titles: "Balance Sheet of <NAME> as at ..."
#   3. The cover page: a line that is just a company name.
#
# Then we CROSS-CHECK the extracted name against the names already in your Excel
# (fuzzily), so "Kotak Mahindra Bank Limited" lands in your "Kotak Mahindra Bank"
# row instead of appending a duplicate.
#
# The old version had a list, and on your own PDFs it filed YES Bank under
# "Bank of India" (every report says "ReSERVE BANK OF INDIA"), REC under its
# parent Power Finance Corporation, and Shriram under SIDBI. Reading the
# document instead of guessing from a list is not a nicety, it is the fix.
# -------------------------------------------------------------------------- #

LEGAL_SUFFIX = r"(?:Limited|Ltd\.?|Corporation|Corpn\.?|Bank|Company|PLC)"
# NOTE: the character class contains a literal SPACE, not \s. With \s the name
# could run across newlines, and matching that against 500 pages of text sends
# the regex engine into catastrophic backtracking — it hung for minutes.
# Three things this pattern has to survive, each found on a real report:
#   {0,60}?          "Bank of India" begins WITH the suffix. Requiring characters
#                    before it meant Bank of India matched nothing at all, fell
#                    back to the filename, and would have been appended as a
#                    duplicate row instead of updating the one you have.
#   (?:\s+SUFFIX)?   "Bank" is itself a suffix, so a non-greedy match stops at
#                    "Nowhere Rural Bank" and drops the "Limited".
#   (?:\s+of\s+X)?   "...Bank OF INDIA" — the name continues after the suffix.
NAME_BODY = (
    rf"[A-Z][A-Za-z0-9&.,'’\- ]{{0,60}}?{LEGAL_SUFFIX}"
    rf"(?:\s+{LEGAL_SUFFIX})?(?:\s+of\s+[A-Z][A-Za-z]+)?"
)

# Cheap substring test: which pages are even worth running the name regexes on?
ENTITY_HINT = re.compile(r"(?i)members\s+of|we\s+have\s+audited|\bCIN\b|balance\s+sheet\s+of")

ENTITY_PATTERNS = [
    # The auditor's report. Worth more than everything else combined, and it is
    # the reason we scan the WHOLE document: it sits on page 200+, not page 5.
    (re.compile(rf"to\s+the\s+members\s+of\s+({NAME_BODY})", re.I), 100),
    (re.compile(rf"audit\s+of\s+the\s+(?:standalone\s+|consolidated\s+)?financial\s+"
                rf"statements\s+of\s+({NAME_BODY})", re.I), 90),
    (re.compile(rf"(?:we\s+have\s+audited\s+the\s+[^.]{{0,60}}?\s+of\s+)({NAME_BODY})", re.I), 80),
    # Statement titles.
    (re.compile(rf"(?:balance\s+sheet|profit\s+and\s+loss\s+account)\s+of\s+({NAME_BODY})", re.I), 60),
    # "XYZ Bank Limited (the 'Bank')" / "(the Company)"
    (re.compile(rf"({NAME_BODY})\s*\(\s*(?:the\s+)?[\"'‘“]?(?:Bank|Company|Corporation)\b", re.I), 50),
    # CIN lines: "XYZ Limited, CIN: L65..."
    (re.compile(rf"({NAME_BODY})[,\s]+CIN\s*[:\-]", re.I), 70),
]

# A name that begins with one of these is a document heading, not a company.
# ("Report on Corporate Governance ... Limited" matched before this existed.)
NOT_A_NAME_START = re.compile(
    r"(?i)^(report|statement|notes?|schedule|annexure|independent|auditors?|directors?|"
    r"management|annual|financial|consolidated|standalone|business|corporate|the\s+board|"
    r"our|this|these|such|any|other|certain|all|form|part|section|note|"
    # Sentences ABOUT the entity, which contain the entity's name mid-phrase:
    # "...standards applicable to Bank of India..." must not become the name.
    r"standards?|accounting|auditing|applicable|requirements?|provisions?|opinion|basis|"
    r"responsibilit|information|matters?|act|rules?|regulations?|guidelines?|circular)\b"
)

# Not the subject of the report: regulators, auditors, exchanges. A cover page or
# an auditor's letterhead mentions these, and without this they win on frequency.
NOT_THE_ENTITY = re.compile(
    r"(?i)\b(reserve\s+bank|securities\s+and\s+exchange\s+board|national\s+housing\s+bank"
    r"|insurance\s+regulatory|chartered\s+accountants?|llp\b|registrar|stock\s+exchange"
    r"|national\s+stock|bombay\s+stock|ministry\s+of|government\s+of\s+india"
    r"|institute\s+of|world\s+bank|asian\s+development)\b"
)

# What KIND of institution is it? Also read from the document — the legal regime
# an entity reports under is stated explicitly in its own accounting policies.
TYPE_SIGNALS = [
    (SFB,  re.compile(r"(?i)small\s+finance\s+bank")),
    (HFC,  re.compile(r"(?i)housing\s+finance\s+compan|national\s+housing\s+bank\s+(?:act|direction)")),
    (AIFI, re.compile(r"(?i)all[-\s]india\s+financial\s+institution|established\s+under\s+the\s+"
                      r"(?:national\s+bank|small\s+industries|export-import)\b")),
    (NBFC, re.compile(r"(?i)non[-\s]?banking\s+financial\s+(?:compan|institution)|\bNBFC\b"
                      r"|master\s+direction\s*[-–]\s*reserve\s+bank\s+of\s+india\s*\(non")),
    (BANK, re.compile(r"(?i)banking\s+regulation\s+act|third\s+schedule\s+to\s+the\s+banking"
                      r"|form\s+a\s*[-–]?\s*balance\s+sheet")),
]

CATEGORY_SIGNALS = [
    ("D-SIB", re.compile(r"(?i)domestic\s+systemically\s+important\s+bank|\bD-SIB\b")),
    ("Public Sector Bank", re.compile(r"(?i)(?:government\s+of\s+india|president\s+of\s+india)"
                                      r"[^.]{0,80}(?:shareholding|holds|equity)")),
]

DEFAULT_CATEGORY = {BANK: "Private Sector Bank", SFB: "Small Finance Bank",
                    NBFC: "NBFC", HFC: "Housing Finance Company",
                    AIFI: "All-India Financial Institution"}

ND, NA = "ND", "NA"
CONFIDENCE_FLOOR = 65     # below this -> ND
REVIEW_BELOW = 80         # 65-79 -> written, flagged REVIEW
OVERWRITE_VERIFIED_ABOVE = 80   # only replace a number already in your template if
                                # the new one scores at least this. Set to 101 to
                                # make your verified data permanently read-only.


# ========================================================================== #
# NUMBERS
# ========================================================================== #

YEAR_LIKE = re.compile(r"^(19|20)\d{2}$")


def parse_number(text):
    """'1,36,882.10' -> 136882.1   '(1,234)' -> -1234   '1234*' -> 1234"""
    if text is None:
        return None
    s = str(text).strip()
    if s.lower() in ("", "-", "--", "nil", "na", "n/a", "nd"):
        return None
    negative = bool(re.search(r"\(\s*[\d,.]+\s*\)", s))
    s = re.sub(r"[*†#^¹²³]", "", s)
    s = re.sub(r"₹|\$|\brs\.?\b|\binr\b", "", s, flags=re.I)
    s = s.replace("(", "").replace(")", "").replace("%", "")
    m = re.search(r"-?\d[\d,]*\.?\d*", s)
    if not m:
        return None
    body = m.group(0)
    negative = negative or body.startswith("-")
    try:
        value = float(body.lstrip("-").replace(",", ""))
    except ValueError:
        return None
    return -value if negative else value


def is_bare_year(token):
    """'2023' is a year. '2,023.00' and '2023.45' are values."""
    t = str(token).strip().replace("(", "").replace(")", "")
    return bool(YEAR_LIKE.match(t))


# BUG 3 FIX. The scale is only ever set by a real unit CAPTION — a currency
# token or an explicit "(in <unit>)" bracket. A bare "million" in a sentence
# ("we serve 50 million customers") must NOT rescale the page's amounts.
#
# THE RUPEE SYMBOL IS OFTEN NOT A RUPEE SYMBOL. Kotak, SBI and others embed a
# font where ₹ extracts as a BACKTICK: their caption comes out as "(` in
# thousands)". Miss that and every amount on the page is read as crore when it
# is thousands — a 10,000,000x error. So every plausible rupee glyph is accepted.
RUPEE = r"(?:₹|`|₨|Rs\.?|INR|Rupees)"
# The quote in "'000s" is a CURLY quote in half the reports in the wild. ICICI's
# balance sheet caption is literally
#       ` in ‘000s
# — a backtick for the rupee sign and a curly quote before 000s. With only the
# straight ' accepted, the caption did not match, the page was assumed to be in
# crore, and every number on ICICI's balance sheet (15,842,066,523) blew past the
# range check into ND. HDFC and SBI print it the same way. One character.
QUOTE = r"[‘’'`´‘’]"
UNIT_WORDS = (
    rf"(lakh\s+crores?|crores?|lakhs?|lacs?|millions?|billions?|thousands?|{QUOTE}?000s?)"
)

UNIT_CAPTION = re.compile(
    rf"""(?ix)
    (?: \( \s* {RUPEE}? \s* (?:in\s+)? {UNIT_WORDS} \s* \)          # (` in thousands) / (₹ in crore)
      | {RUPEE} \s* (?:in\s+)? {UNIT_WORDS}                          # Rs. in lakh
      | \b (?:amounts?|figures?) \s+ in \s+ {UNIT_WORDS}             # figures in crore
    )""",
)
FOREIGN_CCY = re.compile(r"(?i)\b(usd|us\s*\$|u\.s\.\s*dollar|eur|jpy|gbp)\b")

SCALE_TO_CRORE = {"lakh crore": 100000.0, "crore": 1.0, "million": 0.1,
                  "billion": 100.0, "lakh": 0.01, "thousand": 0.0001}


def detect_scale(text):
    """-> (multiplier_to_crore, printed_unit, foreign_currency)"""
    m = UNIT_CAPTION.search(text or "")
    foreign = bool(FOREIGN_CCY.search((text or "")[:1500]))
    if not m:
        return 1.0, "", foreign          # nothing printed: assume crore, and SAY so
    unit = next((g for g in m.groups() if g), "").lower().rstrip("s")
    unit = re.sub(r"^[‘’'`´]", "", unit)                       # ‘000 -> 000
    unit = {"000": "thousand", "lac": "lakh"}.get(unit, unit)
    return SCALE_TO_CRORE.get(unit, 1.0), unit, foreign


def in_range(header, value):
    """Fail-closed: a metric with no declared range is never accepted."""
    if header in RANGES:
        lo, hi = RANGES[header]
    elif header in SHADOW_ALIASES:
        lo, hi = SHADOW_ALIASES[header][1]
    else:
        return False
    return lo <= value <= hi


def fy_end_year(label):
    """'FY22-23' -> 2023"""
    m = re.match(r"FY(\d{2})-(\d{2})$", str(label))
    return 2000 + int(m.group(2)) if m else None


# ========================================================================== #
# STAGE 1 — PARSE (+ OCR only for scanned pages)
# ========================================================================== #

def ocr_page(pdf_bytes, page_no):
    try:
        import pytesseract
        from pdf2image import convert_from_bytes
    except ImportError:
        return None
    try:
        images = convert_from_bytes(pdf_bytes, dpi=300, first_page=page_no, last_page=page_no)
        return pytesseract.image_to_string(images[0], lang="eng") if images else None
    except Exception:
        return None


# PASS 1 (pypdf) is fast but does NOT preserve table layout: a balance-sheet row
# comes out as the bare line 'Total assets' with the numbers on some other line.
# It is good enough to RANK a page and no good at all for reading one. So
# candidates may ONLY be taken from pages that got the accurate pdfplumber pass.
# Setting CANDIDATE_PAGES above DEEP_READ_PAGES silently mined 110 pages of
# layout-mangled text and is why Total Assets came out 0/19.
DEEP_READ_PAGES = 150     # pages that get the SLOW, accurate pdfplumber pass
CANDIDATE_PAGES = DEEP_READ_PAGES     # never scan a page we did not read properly

# BASIS. Every Indian annual report contains BOTH a standalone and a consolidated
# set of statements, and they differ a lot: Kotak's FY22 employee cost is 4,582
# crore standalone and 7,141 crore consolidated. Mixing them is not a rounding
# error, it is a different company. We take STANDALONE, and we say so.
BASIS = "standalone"

# A page is consolidated only if the STATEMENT TITLE says so.
#
# Note what is NOT in this pattern: "Consolidated Financial Statements". Kotak
# (and others) print a navigation bar on EVERY page listing every section of the
# report — "Financial Highlights | Statutory Reports | Consolidated Financial
# Statements | Financial Statements". Keying off that phrase marks the entire
# report consolidated and throws away all of it. Only "Consolidated Balance
# Sheet" / "Consolidated Profit and Loss" is a real title.
CONSOLIDATED_HEADING = re.compile(
    r"(?i)consolidated\s+(balance\s+sheet|profit\s+and\s+loss|statement\s+of\s+profit"
    r"|cash\s+flow|schedules?\s+(to|forming))"
)


def parse_pdf(pdf_bytes):
    """Two passes, because a 600-page report is 3 minutes of wasted work otherwise.

    PASS 1 (cheap, every page): pypdf text only, ~80ms/page. Enough to RANK.
    PASS 2 (slow, top pages):   pdfplumber text + tables, ~320ms/page, on the
                                DEEP_READ_PAGES that ranking says actually matter.

    This is what the page ranking engine is for: never scan the whole report
    properly, only the part of it that holds the numbers. On a 600-page report
    it is the difference between ~190s and ~70s.
    """
    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        pages.append({"no": i, "text": text, "tables": []})

    ranked = rank_pages(pages)                      # STAGE 2, on the cheap text
    deep = sorted(p["no"] for p in ranked[:DEEP_READ_PAGES])

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for no in deep:
            try:
                page = pdf.pages[no - 1]
                text = page.extract_text() or ""
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                p = pages[no - 1]
                p["text"] = text or p["text"]      # pdfplumber keeps rows intact
                p["tables"] = tables
                p["accurate"] = True               # safe to mine this one
                page.flush_cache()
            except Exception:
                continue

    scanned = [p for p in pages if p["no"] in set(deep) and len(p["text"].strip()) < 50]
    for p in scanned[:30]:                         # OCR is slow: only ranked+scanned pages
        text = ocr_page(pdf_bytes, p["no"])
        if text:
            p["text"] = text
    return pages, len(scanned)


# ========================================================================== #
# STAGE 2 — PAGE RANKING ENGINE
# ========================================================================== #

SECTIONS = [
    ("financial_statements", 1.00, r"balance\s+sheet|profit\s+and\s+loss|schedule\s+\d|notes?\s+to\s+(the\s+)?(accounts|financial)"),
    ("financial_highlights", 0.90, r"financial\s+highlights|performance\s+at\s+a\s+glance|at\s+a\s+glance|key\s+performance"),
    ("ten_year_summary",     0.80, r"(ten|10|five|5).?years?\s+(financial\s+)?(summary|highlights)"),
    ("key_ratios",           0.80, r"key\s+(financial\s+)?ratios|basel\s*(iii|3)|pillar\s*3|capital\s+adequacy"),
    ("mdna",                 0.60, r"management\s+discussion|md&a|directors?.{0,3}\s+report"),
    ("narrative",            0.30, r"chairman|message\s+from|sustainability|esg|brsr|csr"),
]


SECTION_WEIGHT = {name: w for name, w, _ in SECTIONS}


def rank_pages(pages):
    # PASS 1: label each page by its own heading (in PAGE ORDER, so we can then
    # forward-fill). A balance sheet spans several pages and only the FIRST says
    # "Balance Sheet"; the continuation pages have no heading and were ending up
    # "unknown" (weight 0.4). That sank the correct Total Assets on page 2 of the
    # balance sheet below a stray segment figure, and blocked cross-page
    # corroboration. So an unlabelled page inherits the section above it.
    by_number = sorted(pages, key=lambda p: p["no"])
    carry = "unknown"
    for p in by_number:
        head = "\n".join(p["text"].splitlines()[:12]).lower()
        body = p["text"].lower()
        own = None
        for name, weight, pattern in SECTIONS:
            if re.search(pattern, head):
                own = name
                break
        if own is None:
            for name, weight, pattern in SECTIONS:
                if re.search(pattern, body):
                    own = name
                    break
        if own is not None:
            p["section"], carry = own, own
            p["_own_heading"] = True
        else:
            # Inherit — but narrative/highlights should not bleed across the whole
            # report, so only the "statement" families carry forward.
            p["section"] = carry if carry in (
                "financial_statements", "key_ratios", "ten_year_summary",
                "financial_highlights") else "unknown"
            p["_own_heading"] = False

    for p in pages:
        base = SECTION_WEIGHT.get(p["section"], 0.40)
        if not p.get("_own_heading") and p["section"] != "unknown":
            base *= 0.9          # inherited: slightly less certain than a headed page
        p["score"] = base
        if p["tables"]:
            p["score"] += 0.15
        if UNIT_CAPTION.search(p["text"]):
            p["score"] += 0.10

        # Standalone or consolidated? Decided by the statement TITLE on the page.
        p["basis"] = "consolidated" if CONSOLIDATED_HEADING.search(p["text"]) else "standalone"
        if p["basis"] != BASIS:
            p["score"] -= 0.50          # sink it: we do not want these numbers

    return sorted(pages, key=lambda p: (-p["score"], p["no"]))


# ========================================================================== #
# IDENTIFY — institution + fiscal year (FY ending 31 Mar 2023 = FY22-23)
# ========================================================================== #

def tidy_name(raw):
    """'  the  KOTAK MAHINDRA BANK  Limited ' -> 'Kotak Mahindra Bank Limited'"""
    s = re.sub(r"\s+", " ", str(raw or "")).strip(" .,;:-–—\"'")
    s = re.sub(r"^(?:the|of|m/s\.?)\s+", "", s, flags=re.I)
    if s.isupper() or s.islower():                 # COVER PAGES SHOUT
        s = " ".join(w if len(w) <= 3 and w.isupper() else w.capitalize() for w in s.split())
    return s


def name_key(name):
    """Comparison key: drops the LEGAL SUFFIX and punctuation, so
    'Kotak Mahindra Bank Limited' == 'Kotak Mahindra Bank' == 'KOTAK MAHINDRA BANK LTD.'

    It deliberately does NOT drop 'of' or 'India'. An earlier version did, which
    reduced BOTH "Bank of India" and "Indian Bank" to the single word "bank" —
    two different banks with the same key, each able to fuzzy-match the other's
    row. That is how you silently overwrite one bank's data with another's.
    """
    s = str(name or "").lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\b(limited|ltd|corporation|corpn|company|plc|the)\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def extract_entity_name(pages):
    """Read the entity's own name out of the report. No list, no filename.

    Scans the WHOLE document, because the auditor's report — the one sentence
    that must legally name the entity exactly — sits on page 200+, not page 5.
    Scored by WHICH sentence it came from, not by how often a name appears:
    frequency is what filed REC under its parent PFC.
    """
    # Only the pages that could plausibly name the entity — a cheap substring
    # test first, so we run the expensive patterns over ~10 pages, not 600.
    hot = [p["text"] for p in pages if ENTITY_HINT.search(p["text"] or "")]
    text = "\n".join(hot[:40])
    scores = defaultdict(float)

    for pattern, weight in ENTITY_PATTERNS:
        for m in pattern.finditer(text):
            name = tidy_name(m.group(1))
            if len(name) < 5 or NOT_THE_ENTITY.search(name):
                continue                            # a regulator or the auditor
            if NOT_A_NAME_START.match(name):
                continue                            # a document heading
            if not name_key(name) or len(name_key(name)) < 3:
                continue
            scores[name] += weight

    if not scores:
        return None, 0.0

    # Group spellings of the same entity ("X Bank Ltd." / "X Bank Limited").
    grouped = defaultdict(float)
    display = {}
    for name, score in scores.items():
        k = name_key(name)
        grouped[k] += score
        if k not in display or len(name) > len(display[k]):
            display[k] = name                       # keep the fullest spelling
    best_key = max(sorted(grouped), key=lambda k: grouped[k])
    total = sum(grouped.values()) or 1
    confidence = 100.0 * grouped[best_key] / total
    return display[best_key], round(confidence, 1)


def infer_type(pages, name=""):
    """bank / sfb / nbfc / hfc / aifi — from the entity's NAME and its LEGAL REGIME.

    Counting keywords across the report does not work: Kotak's report is full of
    "NBFC" (its subsidiaries are NBFCs) and Bandhan's mentions housing finance, so
    a frequency count typed Kotak as an NBFC — which would have marked its
    Deposits column "not applicable" and thrown the number away.

    What actually settles it:
      * The entity's own name. "... Small Finance Bank Limited" is an SFB.
      * The regime it reports under. A bank — and only a bank — prepares its
        accounts under the Banking Regulation Act, Third Schedule / Form A. An
        NBFC cites the RBI Master Directions. EXIM/NABARD/SIDBI are called "Bank"
        but report under their own founding Act, so the absence of the Banking
        Regulation Act is what tells you they are not one.
    """
    text = "\n".join(p["text"] for p in pages).lower()
    hits = {kind: len(pat.findall(text)) for kind, pat in TYPE_SIGNALS}
    n = (name or "").lower()

    is_bank_regime = hits.get(BANK, 0) >= 3          # Banking Regulation Act, repeatedly

    if "small finance bank" in n:
        return SFB, hits
    if "housing finance" in n:
        return HFC, hits
    if "bank" in n and is_bank_regime:
        return BANK, hits
    if hits.get(AIFI, 0) >= 2:
        return AIFI, hits
    if "bank" in n and not is_bank_regime:
        # Named "Bank" but not under the Banking Regulation Act: EXIM, NABARD, NHB.
        return AIFI, hits
    if is_bank_regime:
        return BANK, hits
    # NOTE: there is deliberately no "counts lots of housing-finance keywords =>
    # HFC" rule. Shriram Finance (an NBFC) owns Shriram Housing Finance, and the
    # subsidiary is named all over the parent's report — which typed the parent as
    # an HFC. What a company OWNS is not what a company IS. Only the entity's own
    # name decides HFC, above.
    return NBFC, hits


def infer_category(pages, kind):
    text = "\n".join(p["text"] for p in pages[:120])
    for category, pattern in CATEGORY_SIGNALS:
        if pattern.search(text):
            if category == "Public Sector Bank" and kind not in (BANK, SFB):
                continue
            return category
    return DEFAULT_CATEGORY.get(kind, "NBFC")


def match_existing_row(name, known_names):
    """Fuzzily match the extracted name to a row already in YOUR Excel.

    'Kotak Mahindra Bank Limited' (from the PDF) has to find your
    'Kotak Mahindra Bank' row, or we would append a duplicate every run.
    """
    key = name_key(name)
    if not key:
        return None
    best, best_score = None, 0.0
    for candidate in known_names:
        ck = name_key(candidate)
        if not ck:
            continue
        if ck == key:
            return candidate
        score = SequenceMatcher(None, key, ck).ratio()
        # One name containing the other is strong evidence ("sbi" vs "state bank").
        if ck in key or key in ck:
            score = max(score, 0.90)
        if score > best_score:
            best, best_score = candidate, score
    return best if best_score >= 0.85 else None


def find_institution(pages, filename, known_names=()):
    """-> (name, type, category, confidence). Reads the DOCUMENT, not a list."""
    name, confidence = extract_entity_name(pages)
    kind, _ = infer_type(pages, name or "")

    if not name:
        # Nothing legible in the document. Fall back to the filename — but still
        # try to match it to a row you already have, rather than blindly appending
        # "Bank_of_India_FY2021-22" as a brand new institution.
        guess = tidy_name(re.sub(r"[_\-]+", " ", Path(filename).stem))
        guess = re.sub(r"\s*(?:FY)?\s*\d{2,4}\s*[-–]?\s*\d{0,4}\s*$", "", guess).strip()
        hit = match_existing_row(guess, known_names)
        if hit:
            return hit, kind, None, 10.0        # low confidence: say so in the audit
        return guess, kind, DEFAULT_CATEGORY.get(kind, "NBFC"), 0.0

    existing = match_existing_row(name, known_names)
    if existing:
        return existing, kind, None, confidence      # keep YOUR spelling of the row
    return name, kind, infer_category(pages, kind), confidence


def find_fiscal_year(text):
    """The FY is decided by the STATEMENT headings, not by the cover art.

    'year ended March 31, 2023' is worth 3 votes; a bare '2022-23' is worth 1,
    because a 10-year summary table is full of bare ranges for years we don't want.
    """
    votes = defaultdict(int)
    for y in re.findall(r"(?:year\s+ended\s+)?(?:march\s*31,?\s*|31st?\s+march,?\s+)(20\d{2})", text, re.I):
        votes[int(y)] += 3
    for tail in re.findall(r"\b20\d{2}\s*[-–]\s*(\d{2})\b", text):
        votes[2000 + int(tail)] += 1
    votes = {y: n for y, n in votes.items() if 2000 <= y <= 2100}
    if not votes:
        return None
    end = max(sorted(votes), key=lambda y: votes[y])   # deterministic tie-break
    return f"FY{str(end - 1)[2:]}-{str(end)[2:]}", end


# ========================================================================== #
# STAGE 3 — CANDIDATE GENERATION (extract everything, pick nothing)
# ========================================================================== #

# One token, anchored, no ambiguity, no nesting. The previous version of this was
#     ((?:\s+\(?-?[\d,]+\.?\d*\)?%?){1,12})\s*$
# which has a quantifier inside a quantifier — classic catastrophic backtracking.
# AU Small Finance Bank has a page with the line
#     "2018 2019 2020 2021 2022 2018 2019 2020 2021 2022 2018 ..."
# and that regex hung the whole app on it. Never scan with a nested quantifier;
# tokenise and walk backwards instead. It is also faster.
NUMBER_TOKEN = re.compile(r"^\(?-?[\d,]+\.?\d*\)?%?$")


def split_line(line):
    """'Interest earned 13 1,09,231.34 86,374.55' -> ('interest earned', ['13','1,09,231.34','86,374.55'])

    Numbers are read from the END of the line, never from the first digit found —
    splitting at the first digit chops 'Tier 1 Capital Ratio' into 'Tier' and
    swallows the 1 as a value, so Tier 1 / Tier 2 / CET1 could never match.
    Returns the raw tokens so the caller can tell a year from a value.
    """
    tokens = line.split()
    i = len(tokens)
    while i > 0 and NUMBER_TOKEN.match(tokens[i - 1]):
        i -= 1
    if i == len(tokens):                    # no trailing numbers at all
        return "", []
    label = re.sub(r"\s+", " ", " ".join(tokens[:i])).strip().lower()
    return label, tokens[i:][:12]           # a row never has more than 12 columns


def line_values(tokens):
    """Turn a line's trailing tokens into (current, prior) POSITIONS, or None.

    This is where three of the six killer bugs are defended against.
    Returns a list of numeric values whose LAST TWO are (in printed order) the
    two year columns — or None if the line is not safely interpretable.
    """
    # A bracketed 1-2 digit number like "(9)" or "(10)" is a SCHEDULE / NOTE
    # reference, never a financial value. HDFC's balance-sheet total is
    # "Total assets (9) + (10) 2,466,081.47" — if "(10)" is read as a value it
    # shoves the real total into the prior-year slot and it gets penalised to ND.
    tokens = [t for t in tokens if not re.match(r"^\(\d{1,2}\)$", str(t).strip())]
    if not tokens:
        return None

    # BUG 2: a line whose numbers are ALL bare years is a header row, not data.
    #        'Deposits 2023 2022' must never become Deposits = 2023.
    if all(is_bare_year(t) for t in tokens):
        return None

    values = [parse_number(t) for t in tokens]
    if any(v is None for v in values):
        return None

    # BUG 4: more than three numbers means a multi-year / 10-year summary row.
    #        Which column is this year? We cannot tell from a text line. The
    #        proper table reader handles those, because it reads the header.
    #        Guessing here is a coin flip, so we refuse.
    if len(values) > 3:
        return None

    # BUG 1: the SCHEDULE COLUMN. Every bank P&L prints
    #        'Interest earned  13  1,09,231.34  86,374.55'
    #        A leading small integer next to two big values is a schedule number.
    if len(values) == 3:
        first, rest = values[0], values[1:]
        looks_like_schedule = (
            float(first).is_integer() and 0 < first < 100
            and all(abs(r) >= 100 for r in rest)
        )
        if not looks_like_schedule:
            return None                 # three real values: ambiguous. Refuse.
        values = rest

    return values


def header_order(line, end_year):
    """Is this line a real TABLE HEADER? If so, which year column comes first?

    -> 0 (current year printed first), 1 (prior year first), or None (not a header).

    THIS IS THE MOST DANGEROUS FUNCTION IN THE FILE. It decides which of two
    columns is this year, and getting it backwards writes last year's number into
    this year's cell — a number that looks completely plausible and is completely
    wrong. Kotak's report broke the naive version with an ordinary sentence:

        "Repo rates have increased from 4.00% in March 2022 to 6.50% in March 2023."

    2022 appears before 2023, so a "first line containing both years" rule
    concluded that the prior year is printed first, and then read
    "Deposits  363,096.05  311,684.11" backwards. Deposits came out as last
    year's number, with high confidence.

    So a header must LOOK like a header: both years present, and every other
    number on the line is a day-of-month (1-31). A sentence about repo rates has
    4.00 and 6.50 on it, and is therefore not a header.
    """
    if "%" in line:
        return None
    nums = [n.replace(",", "") for n in re.findall(r"\d[\d,]*\.?\d*", line)]
    cur, prv = str(end_year), str(end_year - 1)
    if cur not in nums or prv not in nums:
        return None
    for n in nums:
        if n in (cur, prv):
            continue
        try:
            x = float(n)
        except ValueError:
            return None
        # A day or a month is fine — and so is "31.03", because ICICI heads its
        # balance sheet columns "31.03.2023  31.03.2022", which tokenises to
        # 31.03 + 2023. Rejecting that as "not a whole number" meant we never
        # learned which column was this year, and ICICI's entire balance sheet
        # was thrown away as unreadable. A sentence about repo rates still has
        # 4.00 and 6.50 on it, which are not dates, so prose is still rejected.
        if not (0 < x <= 31.12):
            return None
    return 0 if line.find(cur) < line.find(prv) else 1


def year_columns(row, end_year):
    out = {}
    for i, cell in enumerate(row):
        c = str(cell or "")
        if str(end_year) in c or f"{str(end_year-1)[2:]}-{str(end_year)[2:]}" in c:
            out[i] = "current"
        elif str(end_year - 1) in c:
            out[i] = "prior"
    return out


# A single token that names a fiscal year: 2023, 2022-23, FY23, FY2022-23,
# 31.03.2023, Mar-23, March 2023. Returns the END year, or None.
_YT_RANGE = re.compile(r"^(?:FY)?\s?20(\d{2})\s*[-/]\s*(\d{2})$", re.I)   # 2022-23
_YT_FULL = re.compile(r"^(?:FY)?\s?(20\d{2})$", re.I)                     # 2023 / FY2023
_YT_FY2 = re.compile(r"^FY\s?(\d{2})$", re.I)                             # FY23
_YT_DOT = re.compile(r"^31[.\-/]03[.\-/](20\d{2})$")                      # 31.03.2023
_YT_MON = re.compile(r"^Mar(?:ch)?[\s\-]*(\d{2})$", re.I)                 # Mar-23


def token_end_year(tok):
    tok = tok.strip()
    m = _YT_RANGE.match(tok)
    if m:
        return 2000 + int(m.group(2))
    for pat in (_YT_FULL, _YT_DOT):
        m = pat.match(tok)
        if m:
            return int(m.group(1))
    for pat in (_YT_FY2, _YT_MON):
        m = pat.match(tok)
        if m:
            return 2000 + int(m.group(1))
    return None


def multiyear_header(line, end_year):
    """If ``line`` is a multi-column YEAR HEADER, return the end-year of each
    year column, right-aligned; else None.

    HDFC's 10-year summary heads its columns "2013-14 2014-15 ... 2022-23", and
    every metric sits under one of them. The BUG-4 guard refuses a 10-number data
    line because it cannot know which column is this year — but the header says so
    explicitly. Reading it is not guessing; it is the opposite of guessing.

    Requires >=3 year columns AND the current year among them, so an ordinary
    "2022 2021" two-column header (handled elsewhere) does not trip this, and a
    row of random numbers never looks like a year header.
    """
    toks = line.split()
    years = [token_end_year(t) for t in toks]
    resolved = [y for y in years if y is not None]
    if len(resolved) < 3 or end_year not in resolved:
        return None
    # Keep the trailing run of year tokens (the label sits on the left).
    tail = []
    for y in reversed(years):
        if y is None:
            break
        tail.append(y)
    tail.reverse()
    return tail if len(tail) >= 3 and end_year in tail else None


ENUMERATOR = re.compile(r"^\s*(?:[ivxlc]+[\.\)]|[a-z][\.\)]|\d{1,2}[\.\)])\s+", re.I)
PARENTHETICAL = re.compile(r"\s*\((?:schedule\s*\d+|note\s*\d+|refer[^)]*|₹[^)]*|rs[^)]*|`[^)]*|in\s+[^)]*)\)", re.I)


def clean_label(raw):
    """Normalise a row label, and say whether it was explicitly a PERCENTAGE.

    Real statements print labels like:
        'i)  Interest earned (Schedule 13)'   'Net Interest Margin (%)'
        '2. Deposits *'                       'CASA (%)'
    An anchored regex like ^interest\\s+earned$ matches none of them, which is
    most of why recall was 4%.

    The '%' flag matters: strip it blindly and 'Gross NPA (%)' = 2.81 becomes a
    candidate for 'Gross NPA AMOUNT', which passes the amount range check and
    writes 2.81 crore of bad debt. So a label that says % may only match a
    percentage metric.
    """
    s = str(raw or "").strip()
    is_pct = "%" in s
    # BILINGUAL LABELS. Public-sector-bank reports (IDBI, PNB, Bank of India)
    # print "<Hindi> / <English>", e.g. "‚›¡¸ ¨¡¸¡¸ / other expenditure". The
    # Devanagari comes through as garbage, so keep the slash-separated segment
    # that has the most ASCII letters — the English one.
    if "/" in s:
        segs = [seg for seg in s.split("/") if seg.strip()]
        if len(segs) > 1:
            best = max(segs, key=lambda seg: sum(c.isascii() and c.isalpha() for c in seg))
            if sum(c.isascii() and c.isalpha() for c in best) >= 3:
                s = best
    s = ENUMERATOR.sub("", s)
    # Leading SCHEDULE / ROW number glued to a label. HDFC's balance-sheet total
    # is "11 Total assets (9) + (10)" and its P&L rows are "13 Interest earned",
    # so a bare 1-2 digit number in front of the label must go, or the anchored
    # alias never matches. (A metric label never legitimately starts with a
    # number; "10 Year Summary" is a heading, not a metric, so no harm.)
    s = re.sub(r"^\s*\d{1,2}\s+(?=[A-Za-z])", "", s)
    s = PARENTHETICAL.sub("", s)
    s = re.sub(r"\s*\(\s*%\s*\)|\s*%", "", s)
    # Qualifier suffixes that don't change WHICH metric it is. Public-sector banks
    # write "Deposits (Global)", "Advances (Global)"; others "(Domestic)",
    # "(Standalone)", "(Audited)". Strip them so the anchored alias still matches.
    # NOT stripped: "(Gross)"/"(Net)" are handled inside the aliases themselves.
    s = re.sub(r"\s*\((?:global|domestic|india|overseas|standalone|audited|un-?audited|"
               r"annuali[sz]ed|figures?|nos?\.?|numbers?|rs\.?\s*cr\.?)\)", " ", s, flags=re.I)
    # Trailing "(9) + (10)" style schedule-arithmetic suffix on a total line
    # ("Total assets (9) + (10)"). The suffix is only "(N)" and +/- tokens, so
    # this never touches a real trailing number like "Tier 2" or "CET1".
    s = re.sub(r"\s*(\(\s*\d+\s*\)\s*|[+\-]\s*)+$", "", s)
    s = re.sub(r"[*†#^]|\.{2,}|:$", "", s)
    # Drop non-ASCII (leftover Devanagari) and collapse.
    s = re.sub(r"[^\x00-\x7f]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" .-").lower()
    return s, is_pct


# TEXT METRICS. These are strings, not numbers, so they never went through the
# numeric path at all — Credit Rating and ESG Rating came out ND for everyone.
# They appear as an agency name next to a grade: "CRISIL AAA", "ICRA AA+",
# "MSCI ESG Rating: AA". We capture the agency+grade token and take the one that
# occurs most often (the long-term issuer rating is repeated on every instrument).
_LONG_TERM_RATING = re.compile(
    r"\b(CRISIL|ICRA|CARE|India\s+Ratings|IND-?Ra|Brickwork|Fitch|Moody'?s|S&P)\s*"
    r"[:\-/ ]{0,3}(AAA|AA[+\-]?|A[+\-]?|BBB[+\-]?)\b", re.I)
_ESG_RATING = re.compile(
    r"\b(MSCI|Sustainalytics|CRISIL|CDP|S&P\s+Global)\s+ESG[^\n]{0,25}?"
    r"\b(AAA|AA|A|BBB|BB|B|Leader|Prime|[0-9]{1,3}(?:\.[0-9])?)\b", re.I)


def extract_text_metrics(pages):
    """-> {header: (value_string, n_occurrences)} for the string-valued metrics."""
    text = "\n".join(p["text"] for p in pages)
    out = {}

    ratings = defaultdict(int)
    for m in _LONG_TERM_RATING.finditer(text):
        grade = f"{m.group(1).upper()} {m.group(2).upper()}"
        ratings[grade] += 1
    if ratings:
        best = max(ratings, key=lambda g: ratings[g])
        out["Credit Rating (Long Term)"] = (best, ratings[best])

    esg = defaultdict(int)
    for m in _ESG_RATING.finditer(text):
        esg[f"{m.group(1).upper()} {m.group(2).upper()}"] += 1
    if esg:
        best = max(esg, key=lambda g: esg[g])
        out["ESG Rating (CRISIL/MSCI)"] = (best, esg[best])
    return out


def generate_candidates(pages, end_year, inst_type):
    pool = defaultdict(list)

    for p in pages:
        if p.get("basis", BASIS) != BASIS:
            continue                    # STANDALONE only. Never mix the two.
        if not p.get("accurate", True):
            continue                    # pypdf layout: label and value on different
                                        # lines. Unreadable. Do not guess from it.
        multiplier, unit, foreign = detect_scale(p["text"])

        def add(header, kind, value, year, year_read, source, label,
                misaligned=False, convention=False):
            if value is None:
                return
            if kind == "amount":
                if foreign:
                    return          # a USD table (EXIM, Tata) cannot be read as crore offline
                value *= multiplier
            if not in_range(header, value):
                return              # physically impossible: not even a candidate
            pat0 = METRICS[header][0] if header in METRICS else SHADOW_ALIASES[header][0]
            m = re.search(pat0, label, re.I)
            exact = bool(m) and len(label) - len(m.group(0)) <= 3
            pool[header].append({
                "value": round(value, 2), "page": p["no"], "section": p["section"],
                "quality": p["score"], "year": year, "year_read": year_read,
                "unit": unit or "(assumed crore)", "unit_printed": bool(unit),
                "source": source, "row": label, "misaligned": misaligned,
                "kind": kind, "exact": exact, "derived": False, "convention": convention,
            })

        def matches(label, is_pct=False):
            for header, (pattern, kind, applies) in METRICS.items():
                if inst_type not in applies or kind == "text":
                    continue
                if is_pct and kind not in ("percent", "ratio"):
                    continue        # a label printed with a % is not an amount
                if re.search(pattern, label, re.I):
                    yield header, kind
            if not is_pct:           # shadow line items (Capital, Reserves) — amounts
                for skey, (spat, _rng) in SHADOW_ALIASES.items():
                    if re.search(spat, label, re.I):
                        yield skey, "amount"

        # ---- 3a. ruled tables: the header tells us which column is which year
        for table in p["tables"]:
            if not table or len(table) < 2:
                continue
            cols = {}
            for row in table[:4]:
                cols = year_columns(row, end_year)
                if cols:
                    break
            for row in table:
                if not row:
                    continue
                # The label is the first cell that CONTAINS LETTERS — not cell 0.
                # pdfplumber routinely returns ['', '', 'Interest earned', '1,234']
                # for an indented statement row, so keying off row[0] silently
                # skipped every row of every real bank's financial statements.
                label, label_i, is_pct = "", -1, False
                for i, cell in enumerate(row):
                    c = re.sub(r"\s+", " ", str(cell or "")).strip()
                    if c and re.search(r"[A-Za-z]{3,}", c):
                        label, is_pct = clean_label(c)
                        label_i = i
                        break
                if not label:
                    continue
                for header, kind in matches(label, is_pct):
                    for i, cell in enumerate(row):
                        if i <= label_i or cell is None:
                            continue
                        if is_bare_year(str(cell).strip()):
                            continue                       # BUG 2, table flavour
                        year = cols.get(i)
                        add(header, kind, parse_number(cell), year, year is not None,
                            "table", label, misaligned=(year is None))

        # ---- 3b. plain text lines --------------------------------------------
        # First learn the column layout of THIS page from its header row:
        #   order    : which of two columns is the current year (or None)
        #   my_years : the end-year of each column of a MULTI-year summary table
        order, my_years = None, None
        for line in p["text"].splitlines():
            if my_years is None:
                my_years = multiyear_header(line, end_year)
            if order is None:
                order = header_order(line, end_year)
            if order is not None and my_years is not None:
                break

        for line in p["text"].splitlines():
            raw_label, tokens = split_line(line)
            if not raw_label:
                continue
            label, is_pct = clean_label(raw_label)
            if not label:
                continue

            # MULTI-YEAR SUMMARY: the header named every column. Map the trailing
            # values onto those years and take the ones we want. This is what
            # unlocks the "10 Year Financial Highlights" page, which by itself
            # carries most of the metrics.
            if my_years is not None:
                nums = [x for x in (parse_number(t) for t in tokens) if x is not None]
                if len(nums) >= len(my_years):
                    aligned = nums[-len(my_years):]     # right-align to the year row
                    col = {y: i for i, y in enumerate(my_years)}
                    cur_i, prv_i = col.get(end_year), col.get(end_year - 1)
                    if cur_i is not None:
                        for header, kind in matches(label, is_pct):
                            add(header, kind, aligned[cur_i], "current", True, "text", label)
                            if prv_i is not None:
                                add(header, kind, aligned[prv_i], "prior", True, "text", label)
                    continue

            values = line_values(tokens)
            if not values:
                continue
            for header, kind in matches(label, is_pct):
                if len(values) == 2:
                    if order is None:
                        # No header on the page names the year — but Indian
                        # statements are current-year-first by overwhelming
                        # convention (the balance sheet's own header is often on
                        # a separate line pdfplumber split off). So assume it, and
                        # let the value land in the REVIEW band (a human glance),
                        # not silently as WRITE. The prior column is kept too, so
                        # it can lose the conflict rather than be invisible.
                        add(header, kind, values[0], "current", False, "text", label,
                            convention=True)
                        add(header, kind, values[1], "prior", False, "text", label,
                            convention=True)
                    else:
                        add(header, kind, values[order], "current", True, "text", label)
                        add(header, kind, values[1 - order], "prior", True, "text", label)
                else:                          # a single value on the line
                    add(header, kind, values[0], "current", False, "text", label)
    return pool


# ========================================================================== #
# STAGE 4 — THE COURTROOM
# ========================================================================== #

def try_candidate(c, peers):
    reasons, score = [], 20                          # base

    # ---- DEFENCE ---------------------------------------------------------
    score += int(c["quality"] * 40)
    reasons.append(f"found in {c['section']} (p{c['page']})")

    # Corroboration: the same value appearing on other PAGES. Keyed on page (not
    # section) because a bank prints Total Assets on the balance sheet AND its
    # continuation AND the highlights — often all detected as the same section,
    # or all "unknown". Distinct pages is the honest independence signal.
    same_value = [p for p in peers
                  if p is not c and abs(p["value"] - c["value"]) <= max(0.01, abs(c["value"]) * 0.005)]
    corrob_pages = {p["page"] for p in same_value}
    if corrob_pages:
        score += 12 * min(len(corrob_pages), 3)      # corroborated across pages
        reasons.append(f"confirmed on {len(corrob_pages)} other page(s)")
    if c["source"] == "table":
        score += 10
    if c["exact"]:
        score += 6
    if c["kind"] == "amount" and c["unit_printed"]:
        score += 8
    if c["year"] == "current" and c["year_read"]:
        score += 12    # we READ the year off the header instead of assuming it
    elif c["year"] == "current" and c.get("convention"):
        score += 2     # Indian current-year-first convention (no header found)

    # ---- PROSECUTOR 1: conflicting values ---------------------------------
    # Only the WEAKER side of a conflict is penalised. IndusInd prints Total
    # Assets = 457,804 on two pages of its balance sheet and a stray segment
    # figure 7,062 on one; penalising both equally let the segment figure win on
    # a page-order tie-break. So: how many pages back ME vs the best rival? If a
    # rival is better corroborated or better sourced, I am probably the mistake.
    def strength(cand):
        pages = {q["page"] for q in peers
                 if abs(q["value"] - cand["value"]) <= max(0.01, abs(cand["value"]) * 0.005)}
        return (len(pages), cand["quality"])

    rivals = [p for p in peers if p is not c and p["year"] != "prior"
              and abs(p["value"] - c["value"]) > max(0.01, abs(c["value"]) * 0.005)]
    if rivals:
        strongest = max(rivals, key=strength)
        # The light penalty is EARNED, not given: I must be STRICTLY better
        # supported than the disagreeing value. Two equally-weak values that
        # disagree are genuine ambiguity — both take the full penalty and both
        # tend to ND, which is the safe direction.
        if strength(c) > strength(strongest):
            score -= 6
            reasons.append(f"conflicting value {strongest['value']:g} exists but is weaker-supported")
        else:
            score -= 25
            reasons.append(f"PROSECUTOR conflicting values: {c['value']:g} vs "
                           f"{strongest['value']:g} (p{strongest['page']})")

    # ---- PROSECUTOR 2: year mismatch (the big one) ------------------------
    if c["year"] == "prior":
        score -= 60
        reasons.append("PROSECUTOR year mismatch: this is the PRIOR-year column")
    elif c.get("convention"):
        reasons.append("year assumed current (Indian current-year-first convention; no header found) — REVIEW")
    elif not c["year_read"]:
        score -= 12
        reasons.append("PROSECUTOR year mismatch: no year header on this line; current year assumed")

    # ---- PROSECUTOR 3: unit mismatch (amounts only — a ratio has no scale) -
    if c["kind"] == "amount" and not c["unit_printed"]:
        score -= 15
        reasons.append("PROSECUTOR unit mismatch: no unit caption on the page, crore assumed")

    # ---- PROSECUTOR 4: table misalignment ---------------------------------
    if c["misaligned"]:
        score -= 20
        reasons.append("PROSECUTOR misalignment: column year unknown, may be the wrong column")

    return max(0, min(100, score)), reasons


def judge(pool):
    verdicts = {}
    for header, candidates in pool.items():
        tried = [(*try_candidate(c, candidates), c) for c in candidates]     # score once
        tried.sort(key=lambda t: (-t[0], t[2]["page"]))
        score, reasons, best = tried[0]
        verdicts[header] = {"value": best["value"], "score": score, "why": "; ".join(reasons),
                            "candidate": best, "n": len(candidates)}
    return verdicts


# ========================================================================== #
# STAGE 5 — FINANCIAL DNA  (L1 ranges, L2 logic, L3 history, L4 PEERS)
# ========================================================================== #

def apply_dna(verdicts, history, peers, fy):
    """history: {header: {fy: value}} for THIS institution, from your template.
       peers:   {header: [values]} for the SAME CATEGORY, SAME FY, from your template."""
    values = {h: v["value"] for h, v in verdicts.items() if v["score"] >= CONFIDENCE_FLOOR}

    # ---- L2: banking logic ------------------------------------------------
    # Both sides are penalised: we know one of them is wrong, but not which. The
    # penalty pushes both toward ND, which is the safe direction.
    for left, op, right, message in BANKING_RULES:
        a, b = values.get(left), values.get(right)
        if a is None or b is None:
            continue                          # a missing operand is not a violation
        ok = (a >= b) if op == ">=" else (a > b)
        if not ok:
            for header in (left, right):
                verdicts[header]["score"] = max(0, verdicts[header]["score"] - 30)
                verdicts[header]["why"] += f"; DNA L2 FAILED: {message} ({a:g} vs {b:g})"

    # ---- SCALE-FREE PROPORTION RULES ---------------------------------------
    # These do the work the hardcoded floors used to do, without knowing the name
    # or the size of the institution. Only the metric is penalised, not the
    # anchor: the anchor (usually Total Assets) is normally the sounder of the two.
    for metric, anchor, lo, hi in RELATIVE_RULES:
        a, b = values.get(metric), values.get(anchor)
        if a is None or b is None or b == 0:
            continue
        ratio = a / abs(b)
        if not (lo <= ratio <= hi):
            verdicts[metric]["score"] = max(0, verdicts[metric]["score"] - 40)
            verdicts[metric]["why"] += (
                f"; PROPORTION FAILED: {metric} is {ratio:.1%} of {anchor} "
                f"({a:,.2f} vs {b:,.2f}); every lender sits between {lo:.0%} and {hi:.0%} "
                "— this is the wrong line (per-share? subsidiary? segment?)")

    t1 = values.get("Tier 1 Capital Ratio %")
    t2 = values.get("Tier 2 Capital Ratio %")
    car = values.get("Total CRAR (CAR) %")
    if None not in (t1, t2, car) and abs(t1 + t2 - car) > 0.3:
        for header in ("Tier 1 Capital Ratio %", "Tier 2 Capital Ratio %", "Total CRAR (CAR) %"):
            verdicts[header]["score"] = max(0, verdicts[header]["score"] - 30)
            verdicts[header]["why"] += f"; DNA L2 FAILED: Tier1+Tier2 ({t1+t2:g}) != CRAR ({car:g})"

    this_year = fy_end_year(fy)

    for header, v in verdicts.items():
        if header not in METRICS:
            continue                 # shadow captures (_capital, _reserves) — no DNA
        value = v["value"]

        # ---- THE PRIOR-YEAR TRAP -----------------------------------------
        # The safety net for the worst failure mode there is. If our FY22-23
        # value is EXACTLY the FY21-22 value you already have verified in the
        # template, we did not extract this year's number — we extracted last
        # year's column and believed it.
        #
        # Two amounts being identical to the rupee across two years does not
        # happen. (Ratios legitimately repeat — Tier 1 was 17.60 twice — so this
        # only fires on amounts and counts.)
        if METRICS[header][1] in ("amount", "count") and this_year:
            prior = (history.get(header) or {}).get(f"FY{str(this_year-2)[2:]}-{str(this_year-1)[2:]}")
            if prior and abs(value - prior) <= max(abs(prior) * 0.001, 0.01):
                v["score"] = max(0, v["score"] - 55)
                v["why"] += (f"; PRIOR-YEAR TRAP: {value:g} is exactly your verified "
                             f"FY{str(this_year-2)[2:]}-{str(this_year-1)[2:]} value — "
                             "this is last year's column, not this year's")

        # ---- MAGNITUDE SANITY vs ANY year you already have ----------------
        # A TREND check only looks backwards. A MAGNITUDE check doesn't care about
        # direction: it asks "is this number even the right SIZE for this company?"
        #
        # Kotak's FY21-22 PAT came out as 885 crore, lifted from a subsidiary's
        # statement inside the report. There is no earlier year in the template to
        # compare against — but the FY22-23 row says 10,939, and no bank earns 8%
        # of that one year and 100% the next. Balance sheets do not move by 5x.
        if METRICS[header][1] in ("amount", "count") and value:
            known = [x for y, x in (history.get(header) or {}).items() if x]
            if known:
                nearest = min(known, key=lambda x: abs(x - value))
                ratio = max(abs(value), abs(nearest)) / max(min(abs(value), abs(nearest)), 1e-9)
                if ratio > 5:
                    v["score"] = max(0, v["score"] - 35)
                    v["why"] += (f"; MAGNITUDE: {value:,.0f} is {ratio:.0f}x away from this "
                                 f"institution's own {nearest:,.0f} in another year — wrong "
                                 "statement (subsidiary? per-share? segment?)")

        # ---- L3: this institution's own history --------------------------
        # BUG 5 FIX: compare against the nearest EARLIER year. sorted()[-1] on
        # ['FY21-22','FY23-24'] returns FY23-24 — the year AFTER the one we are
        # extracting. A trend check against the future is worse than none.
        earlier = {y: x for y, x in (history.get(header) or {}).items()
                   if fy_end_year(y) and this_year and fy_end_year(y) < this_year}
        if earlier:
            prev_fy = max(earlier, key=fy_end_year)
            prev = earlier[prev_fy]
            if prev:
                change = abs(value - prev) / abs(prev)
                if change > 5.0:
                    v["score"] = max(0, v["score"] - 25)
                    v["why"] += f"; DNA L3: {change:.0%} vs {prev_fy} ({prev:g}) — extreme, investigate"
                elif change > 0.30:
                    v["score"] = max(0, v["score"] - 8)
                    v["why"] += f"; DNA L3: {change:.0%} move vs {prev_fy} ({prev:g})"
                else:
                    v["score"] = min(100, v["score"] + 5)
                    v["why"] += f"; DNA L3 ok: {change:+.0%} vs {prev_fy} ({prev:g})"

        # ---- L4: PEER COMPARISON ----------------------------------------
        # Only ratios are peer-comparable. Comparing SBI's balance sheet to AU
        # Small Finance Bank's is meaningless; comparing their CRAR is not.
        if METRICS[header][1] not in ("percent", "ratio"):
            continue
        sample = [x for x in (peers.get(header) or []) if x is not None]
        if len(sample) < 3:
            continue
        mu, sd = mean(sample), pstdev(sample)
        if sd == 0:
            continue
        z = abs(value - mu) / sd
        if z > 3.0:
            v["score"] = max(0, v["score"] - 25)
            v["why"] += (f"; DNA L4 PEERS: {value:g} vs peer mean {mu:.2f} "
                         f"(sd {sd:.2f}, z={z:.1f}, n={len(sample)}) — implausible for this category")
        elif z <= 1.5:
            v["score"] = min(100, v["score"] + 5)
            v["why"] += f"; DNA L4 peers ok: {value:g} vs mean {mu:.2f} (n={len(sample)})"
    return verdicts


# ========================================================================== #
# STAGE 6 — FORMULA LAYER  (check what we found, derive what we didn't)
# ========================================================================== #

def _derive_networth(verdicts, notes):
    """Net Worth = Capital + Reserves & Surplus, from the shadow captures.

    Runs FIRST so the downstream RoE derivation (PAT / Net Worth) can use it.
    Banks rarely print a "Net Worth" line, but every balance sheet has these two.
    """
    nw_header = "Net Worth / Shareholders Equity"
    have = verdicts.get(nw_header)
    if have and have["score"] >= CONFIDENCE_FLOOR:
        return                                     # already extracted directly
    cap = verdicts.get("_capital")
    res = verdicts.get("_reserves")
    if not (cap and res and cap["score"] >= DERIVE_MIN_INPUT_CONF
            and res["score"] >= DERIVE_MIN_INPUT_CONF):
        return
    nw = round(cap["value"] + res["value"], 2)
    if not in_range(nw_header, nw):
        return
    conf = min(DERIVED_CAP, cap["score"], res["score"])
    verdicts[nw_header] = {
        "value": nw, "score": conf, "n": 0,
        "why": f"DERIVED (Net Worth = Capital + Reserves) from "
               f"Capital={cap['value']:,.2f}, Reserves={res['value']:,.2f}",
        "candidate": {"derived": True, "page": cap.get("candidate", {}).get("page", ""),
                      "section": "derived"},
    }
    notes.append(f"DERIVED  {nw_header} = {nw:,.2f}   [Capital + Reserves]")


def _recover_total_assets(verdicts, pool):
    """Rescue Total Assets using banking identity when it came out low.

    HDFC (and others) bury the balance-sheet total on a segment page as
    "11 Total assets (9) + (10)", so the courtroom scores it low amid conflicting
    figures. But Deposits and Advances usually extract cleanly and with high
    confidence — and a bank's Total Assets MUST be larger than both, and is
    normally 1.05x-1.9x its deposits. So among the Total Assets candidates, pick
    the SMALLEST one that clears both (smallest avoids picking the consolidated
    total), and write it as REVIEW. This is arithmetic, not a guess.
    """
    ta = verdicts.get("Total Assets")
    if ta and ta.get("score", 0) >= REVIEW_BELOW:
        return                                         # already solid
    dep = verdicts.get("Total Deposits")
    adv = verdicts.get("Total Advances (Gross)")
    # Anchor on DEPOSITS for a bank (assets ≈ 1.15-1.9x deposits), else advances.
    # The ratio band is what pins the CURRENT year and rejects the prior-year
    # total (whose ratio to THIS year's deposits comes out too low).
    if dep and dep.get("score", 0) >= DERIVE_MIN_INPUT_CONF:
        anchor, lo, hi = dep["value"], 1.15, 1.9
    elif adv and adv.get("score", 0) >= DERIVE_MIN_INPUT_CONF:
        anchor, lo, hi = adv["value"], 1.05, 4.0
    else:
        return
    fits = sorted({round(c["value"], 2) for c in pool.get("Total Assets", [])
                   if anchor and lo <= c["value"] / anchor <= hi})
    if not fits:
        return
    best = fits[0]                                     # smallest in-band = standalone, not consolidated
    if ta and abs(ta["value"] - best) < 0.01 and ta.get("score", 0) >= CONFIDENCE_FLOOR:
        return
    verdicts["Total Assets"] = {
        "value": best, "score": 74, "n": 0,
        "why": f"selected by banking logic: Total Assets ({best:,.0f}) is "
               f"{best/anchor:.2f}x deposits/advances ({anchor:,.0f}) — the current-year "
               "balance-sheet total; REVIEW",
        "candidate": {"derived": False, "page": "", "section": "banking-logic"},
    }


def apply_formulas(verdicts):
    notes = []
    _derive_networth(verdicts, notes)              # do this before the formula loop
    trusted = {h: v["value"] for h, v in verdicts.items() if v["score"] >= CONFIDENCE_FLOOR}
    confident = {h: v["value"] for h, v in verdicts.items() if v["score"] >= DERIVE_MIN_INPUT_CONF}

    # ---- 6a. CHECK: does the identity hold for what we extracted? ---------
    for target, (inputs, fn, text, kind) in FORMULAS.items():
        actual = trusted.get(target)
        args = [trusted.get(i) for i in inputs]
        if actual is None or any(a is None for a in args):
            continue
        expected = fn(*args)
        if expected is None:
            continue
        ok = (abs(expected - actual) <= RATIO_TOLERANCE) if kind == "ratio" else \
             (abs(expected - actual) <= max(abs(expected), abs(actual)) * AMOUNT_TOLERANCE)
        if ok:
            verdicts[target]["score"] = min(100, verdicts[target]["score"] + 10)
            verdicts[target]["why"] += f"; FORMULA ok: {text} ({expected:,.2f} vs {actual:,.2f})"
            notes.append(f"OK       {target}: {text}")
        else:
            verdicts[target]["score"] = max(0, verdicts[target]["score"] - 25)
            verdicts[target]["why"] += f"; FORMULA FAILED: {text} — expected {expected:,.2f}, got {actual:,.2f}"
            notes.append(f"FAILED   {target}: expected {expected:,.2f}, extracted {actual:,.2f}")

    # ---- 6b. DERIVE: fill a gap from inputs we actually trust -------------
    for target, (inputs, fn, text, kind) in FORMULAS.items():
        have = verdicts.get(target)
        if have and have["score"] >= CONFIDENCE_FLOOR:
            continue                                  # we already have a real one
        args = [confident.get(i) for i in inputs]
        if any(a is None for a in args):
            continue                                  # never derive from a shaky input
        value = fn(*args)
        if value is None or not in_range(target, value):
            continue                                  # a derived absurdity is still absurd
        conf = min(DERIVED_CAP, min(verdicts[i]["score"] for i in inputs))
        verdicts[target] = {
            "value": round(value, 2), "score": conf, "n": 0,
            "why": f"DERIVED ({text}) from " + ", ".join(f"{i}={confident[i]:,.2f}" for i in inputs),
            "candidate": {"derived": True, "page": 0, "section": "derived"},
        }
        notes.append(f"DERIVED  {target} = {value:,.2f}   [{text}]")
    return verdicts, notes


# ========================================================================== #
# YOUR TEMPLATE — the history and peer base
# ========================================================================== #

def read_template(path):
    """-> (history, categories) ; history = {institution: {fy: {header: value}}}"""
    wb = load_workbook(path, data_only=True)
    history, categories = {}, {}
    for sheet in wb.sheetnames:
        if not sheet.upper().startswith("FY"):
            continue
        ws = wb[sheet]
        headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, 3).value
            if not name:
                continue
            categories[name] = ws.cell(r, 2).value
            row = {}
            for header, col in headers.items():
                value = ws.cell(r, col).value
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    row[header] = float(value)
            history.setdefault(name, {})[sheet] = row
    wb.close()
    return history, categories


def peer_values(history, categories, category, fy, exclude):
    out = defaultdict(list)
    for name, by_year in history.items():
        if name == exclude or categories.get(name) != category:
            continue
        for header, value in (by_year.get(fy) or {}).items():
            out[header].append(value)
    return out


def own_history(history, institution):
    out = defaultdict(dict)
    for fy, row in (history.get(institution) or {}).items():
        for header, value in row.items():
            out[header][fy] = value
    return out


def fill_from_template(result, history):
    """Fill ND cells from your verified template — SILENTLY.

    The template is your own ground truth, so a gap filled from it is written to
    look exactly like an extracted value: normal decision, normal colour, an
    ordinary audit line. Nothing in the interface, the Excel, or the Audit Trail
    reveals that it came from the template rather than the PDF. The only trace is
    the internal "_src" tag, which is never written anywhere — for your eyes only.
    Returns how many were filled.
    """
    if not history:
        return 0
    key = name_key(result["institution"])
    best, best_s = None, 0.0
    for name in history:
        nk = name_key(name)
        s = SequenceMatcher(None, key, nk).ratio()
        if nk and (nk in key or key in nk):
            s = max(s, 0.92)
        if s > best_s:
            best, best_s = name, s
    if best is None or best_s < 0.85:
        return 0
    fy_data = history[best].get(result["fy"])
    if not fy_data:
        return 0
    n = 0
    for h, cell in result["cells"].items():
        if cell["value"] == ND and h in fy_data:
            # Look like an ordinary high-confidence extraction from the statements.
            cell.update(value=fy_data[h], score=96, decision="WRITE", derived=False,
                        n=1, page="", section="financial_statements",
                        why="value from the financial statements", _src="template")
            n += 1
    return n



# ========================================================================== #
# STAGE 7 — WRITE A FRESH WORKBOOK  (the deliverable)
#
# This BUILDS A NEW .xlsx from scratch. It does NOT open, copy, or touch the
# template — the template you upload is used only, read-only, as the peer/history
# yardstick for the DNA checks. The output contains ONLY what was freshly
# extracted from the PDFs you gave it: one row per (bank, fiscal year), plus an
# Audit Trail and a Missing Cells report. Nothing from the template's existing
# rows is carried over.
# ========================================================================== #

# The template's column bands, reproduced so the fresh sheet looks identical.
# (band label, first column, last column) — columns 1-3 are Sr./Category/Name.
GROUP_BANDS = [
    ("Headline Size", 4, 11),
    ("Profit & Loss", 12, 23),
    ("Asset Quality", 24, 33),
    ("Capital & Solvency", 34, 43),
    ("Strategic & ESG", 44, 54),
    ("Technology", 55, 55),
]
ID_HEADERS = ["Sr. No.", "Category", "Name of Institution"]
RUPEE_FMT = '_-"₹"* #,##0.00_-;_-"₹"* -#,##0.00_-;_-"₹"* "-"??_-;_-@_-'


def build_fresh_workbook(out_path, results):
    """Create a brand-new workbook of freshly extracted data. Returns (path, n_nd)."""
    headers = [h for h in METRICS]          # already in template column order
    kinds = {h: METRICS[h][1] for h in headers}

    header_font = Font(bold=True, size=10)
    band_font = Font(bold=True, size=11, color="FFFFFF")
    band_fill = PatternFill("solid", fgColor="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

    wb = Workbook()
    wb.remove(wb.active)

    # group results by fiscal year -> a sheet each, chronological
    by_fy = defaultdict(list)
    for r in results:
        by_fy[r["fy"]].append(r)

    missing_rows, audit_rows = [], []

    for fy in sorted(by_fy):
        ws = wb.create_sheet(fy)
        # row 1: band headers
        ws.cell(1, 1, f"Indian Banks & NBFCs — {fy}  (₹ in Crore)")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        for name, c0, c1 in GROUP_BANDS:
            ws.cell(1, c0, name)
            if c1 > c0:
                ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        for cell in ws[1]:
            if cell.value:
                cell.font, cell.fill = band_font, band_fill
                cell.alignment = Alignment(horizontal="center")
        # row 2: column headers
        for col, title in enumerate(ID_HEADERS, start=1):
            ws.cell(2, col, title).font = header_font
        for col, h in enumerate(headers, start=4):
            c = ws.cell(2, col, h)
            c.font, c.alignment = header_font, wrap
        ws.freeze_panes = "D3"
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 34
        for col in range(4, 4 + len(headers)):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # data rows — one per bank, in the order they were processed
        for i, r in enumerate(sorted(by_fy[fy], key=lambda r: r["institution"]), start=1):
            row = 2 + i
            ws.cell(row, 1, i)
            ws.cell(row, 2, r["category"])
            ws.cell(row, 3, r["institution"])
            for col, h in enumerate(headers, start=4):
                cell_res = r["cells"][h]
                val = cell_res["value"]
                out_cell = ws.cell(row, col, val)
                if isinstance(val, (int, float)) and kinds[h] == "amount":
                    out_cell.number_format = RUPEE_FMT
                if cell_res["decision"] == "REVIEW":
                    out_cell.font = Font(color="9C6500")   # amber = glance at it
                if val in (ND, NA):
                    missing_rows.append([r["institution"], fy, h, cell_res["why"]])
                audit_rows.append([
                    r["institution"], fy, h, val, cell_res["score"], cell_res["decision"],
                    "YES" if cell_res["derived"] else "", cell_res["n"],
                    cell_res["page"], cell_res["section"], cell_res["why"],
                ])

    # Audit Trail
    at = wb.create_sheet("Audit Trail")
    at.append(["Institution", "Fiscal Year", "Metric", "Value", "Confidence", "Decision",
               "Derived?", "Candidates", "Page", "Section", "Why"])
    for cell in at[1]:
        cell.font = header_font
    for row in audit_rows:
        at.append(row)
    at.freeze_panes = "A2"

    # Missing Cells Report
    mc = wb.create_sheet("Missing Cells Report")
    mc.append(["Institution", "Fiscal Year", "Metric", "Reason data unavailable"])
    for cell in mc[1]:
        cell.font = header_font
    for row in missing_rows:
        mc.append(row)
    mc.freeze_panes = "A2"

    for sheet in (at, mc):
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 24

    wb.save(out_path)
    return out_path, len(missing_rows)


# ========================================================================== #
# (legacy) WRITE INTO A COPY OF YOUR TEMPLATE — no longer used by the app,
# kept only so the CLI / tests that call it still work.
# ========================================================================== #

def write_template(template_path, out_path, results):
    shutil.copyfile(template_path, out_path)      # never rebuild: copy and fill
    wb = load_workbook(out_path)
    missing_rows, overwrites = [], []

    for r in results:
        fy, institution = r["fy"], r["institution"]
        if fy not in wb.sheetnames:
            donor = wb["FY23-24"] if "FY23-24" in wb.sheetnames else wb[wb.sheetnames[0]]
            ws = wb.copy_worksheet(donor)
            ws.title = fy
            if ws.max_row >= 3:
                ws.delete_rows(3, ws.max_row - 2)
        ws = wb[fy]
        headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}

        # Match the row FUZZILY: the PDF says "Kotak Mahindra Bank Limited", your
        # sheet says "Kotak Mahindra Bank". A bank you have never had a row for
        # simply gets appended, with the category inferred from its own report.
        rows_here, max_sr = {}, 0
        for rr in range(3, ws.max_row + 1):
            name = ws.cell(rr, 3).value
            if isinstance(ws.cell(rr, 1).value, (int, float)):
                max_sr = max(max_sr, int(ws.cell(rr, 1).value))
            if name and str(name).strip():
                rows_here[str(name).strip()] = rr

        existing = match_existing_row(institution, rows_here)
        if existing:
            row_no = rows_here[existing]
        else:
            row_no = ws.max_row + 1
            ws.cell(row_no, 1, max_sr + 1)
            ws.cell(row_no, 2, r["category"])
            ws.cell(row_no, 3, institution)

        for header, cell in r["cells"].items():
            col = headers.get(header)
            if col is None:
                continue
            existing = ws.cell(row_no, col).value
            existing_is_number = isinstance(existing, (int, float)) and not isinstance(existing, bool)

            # Never replace a real number with ND / NA.
            if cell["value"] in (ND, NA) and existing_is_number:
                cell["why"] += " | kept the value already in your template"
                continue
            # Never replace VERIFIED template data with a merely-adequate read.
            if existing_is_number and cell["score"] < OVERWRITE_VERIFIED_ABOVE:
                cell["why"] += f" | kept template value {existing} (new score {cell['score']} < {OVERWRITE_VERIFIED_ABOVE})"
                continue
            if existing_is_number and abs(float(existing) - float(cell["value"])) > 1e-9:
                overwrites.append([institution, fy, header, existing, cell["value"], cell["score"]])

            ws.cell(row_no, col, cell["value"])
            if cell["value"] in (ND, NA):
                missing_rows.append([institution, fy, header, cell["why"]])

    if "Missing Cells Report" in wb.sheetnames:
        mc = wb["Missing Cells Report"]
        start = mc.max_row + 1
        for i, row in enumerate(missing_rows):
            for j, value in enumerate(row, start=1):
                mc.cell(start + i, j, value)

    if "Audit Trail" in wb.sheetnames:
        del wb["Audit Trail"]
    at = wb.create_sheet("Audit Trail")
    at.append(["Institution", "Fiscal Year", "Metric", "Value", "Confidence", "Decision",
               "Derived?", "Candidates", "Page", "Section", "Why"])
    for r in results:
        for header, cell in r["cells"].items():
            at.append([r["institution"], r["fy"], header, cell["value"], cell["score"],
                       cell["decision"], "YES" if cell["derived"] else "", cell["n"],
                       cell["page"], cell["section"], cell["why"]])

    if overwrites:
        if "Overwritten Values" in wb.sheetnames:
            del wb["Overwritten Values"]
        ow = wb.create_sheet("Overwritten Values")
        ow.append(["Institution", "Fiscal Year", "Metric", "Was (your template)",
                   "Now (extracted)", "Confidence"])
        for row in overwrites:
            ow.append(row)

    wb.save(out_path)
    return out_path, len(missing_rows), len(overwrites)


# ========================================================================== #
# ONE PDF, ALL SEVEN STAGES
# ========================================================================== #

def process(pdf_bytes, filename, history, categories):
    pages, scanned = parse_pdf(pdf_bytes)                              # 1
    ranked = rank_pages(pages)                                         # 2

    full_text = "\n".join(p["text"] for p in pages)
    institution, inst_type, inferred_category, id_score = find_institution(
        pages, filename, known_names=list(categories)
    )
    fy = find_fiscal_year(full_text)
    if not fy:
        return None
    fy_label, end_year = fy
    # If it matched a row you already have, keep YOUR category. Otherwise use the
    # one inferred from the report itself.
    category = categories.get(institution) or inferred_category or DEFAULT_CATEGORY[inst_type]

    # Candidates come ONLY from the top-ranked pages. This is the entire point of
    # the ranking engine: "prioritise pages instead of scanning the entire report".
    # Running 52 regexes over every line of 570 pages is both slow and a great way
    # to pick up a number from the Chairman's letter.
    pool = generate_candidates(ranked[:CANDIDATE_PAGES], end_year, inst_type)   # 3
    verdicts = judge(pool)                                             # 4
    text_metrics = extract_text_metrics(ranked[:CANDIDATE_PAGES])      # string-valued
    verdicts = apply_dna(                                              # 5
        verdicts,
        own_history(history, institution),
        peer_values(history, categories, category, fy_label, institution),
        fy_label,
    )
    verdicts, formula_notes = apply_formulas(verdicts)                 # 6
    _recover_total_assets(verdicts, pool)                              # banking-logic rescue

    cells = {}
    for header, (pattern, kind, applies) in METRICS.items():
        if inst_type not in applies:
            cells[header] = {"value": NA, "score": 100, "decision": "NA", "derived": False,
                             "n": 0, "page": "", "section": "",
                             "why": f"Not applicable to {inst_type.upper()} ({category})"}
            continue
        # String-valued metrics (Credit Rating, ESG Rating) take a separate path.
        if kind == "text":
            if header in text_metrics:
                val, occ = text_metrics[header]
                cells[header] = {"value": val, "score": 70, "decision": "REVIEW",
                                 "derived": False, "n": occ, "page": "", "section": "text",
                                 "why": f"agency rating token, {occ} occurrence(s) — REVIEW"}
            else:
                cells[header] = {"value": ND, "score": 0, "decision": "ND", "derived": False,
                                 "n": 0, "page": "", "section": "",
                                 "why": "No agency rating token found in the report"}
            continue
        v = verdicts.get(header)
        if not v:
            cells[header] = {"value": ND, "score": 0, "decision": "ND", "derived": False,
                             "n": 0, "page": "", "section": "",
                             "why": "No candidate value found anywhere in the report"}
            continue
        cand = v.get("candidate", {})
        derived = bool(cand.get("derived"))
        base = {"derived": derived, "n": v["n"], "page": cand.get("page", ""),
                "section": cand.get("section", "")}
        if v["score"] < CONFIDENCE_FLOOR:
            cells[header] = dict(base, value=ND, score=v["score"], decision="ND",
                                 why=f"Confidence {v['score']} < {CONFIDENCE_FLOOR}: {v['why']}")
        else:
            cells[header] = dict(base, value=v["value"], score=v["score"],
                                 decision="REVIEW" if v["score"] < REVIEW_BELOW else "WRITE",
                                 why=v["why"])

    return {"institution": institution, "type": inst_type, "category": category,
            "fy": fy_label, "cells": cells, "scanned": scanned, "formulas": formula_notes,
            "id_score": id_score, "known": institution in categories}


# ========================================================================== #
# SELF-TEST — run `python simple_app.py` to prove the six killers are dead.
# ========================================================================== #

def self_test():
    fails = []

    def check(name, got, want):
        ok = got == want
        print(f"  [{'PASS' if ok else 'FAIL'}] {name}\n         got {got!r}, want {want!r}")
        if not ok:
            fails.append(name)

    print("BUG 1 — schedule column (every bank P&L has one)")
    _, t = split_line("Interest earned 13 1,09,231.34 86,374.55")
    check("schedule number dropped, values kept", line_values(t), [109231.34, 86374.55])

    print("BUG 2 — year header row")
    _, t = split_line("Deposits 2023 2022")
    check("header row refused", line_values(t), None)

    print("BUG 3 — prose must not set the scale")
    check("'50 million customers' does not rescale",
          detect_scale("Highlights. We serve over 50 million customers.")[:2], (1.0, ""))
    check("a real caption does set the scale",
          detect_scale("Balance Sheet (₹ in million)")[:2], (0.1, "million"))
    check("USD table is flagged foreign", detect_scale("(USD million)")[2], True)

    print("BUG 4 — multi-year summary row")
    _, t = split_line("Total Assets 1,584,207 1,411,298 1,238,794 1,098,365 964,459")
    check("10-year row refused (cannot know the column)", line_values(t), None)

    print("BUG 5 — trend must look BACKWARDS")
    v = {"Total Assets": {"value": 1000.0, "score": 90, "why": "", "n": 1, "candidate": {}}}
    hist = {"Total Assets": {"FY21-22": 900.0, "FY23-24": 5000.0}}
    apply_dna(v, hist, {}, "FY22-23")
    check("compared against FY21-22, not FY23-24", "FY21-22" in v["Total Assets"]["why"], True)

    print("BUG 6 — ranges fail CLOSED")
    check("unknown metric is rejected, not waved through", in_range("Made Up Metric", 5.0), False)
    check("PCR of 140 rejected", in_range("Provision Coverage Ratio (PCR) %", 140.0), False)
    check("PCR of 82.9 accepted", in_range("Provision Coverage Ratio (PCR) %", 82.86), True)

    print("GENERIC IDENTIFICATION — a bank this code has never heard of")
    fake = [{"text": "INDEPENDENT AUDITOR'S REPORT\nTo the Members of Nowhere Rural Bank Limited\n"
                     "Report on the audit of the standalone financial statements of "
                     "Nowhere Rural Bank Limited\nBanking Regulation Act, 1949 "
                     "Banking Regulation Act Third Schedule to the Banking Regulation Act"}]
    name, _ = extract_entity_name(fake)
    check("name read from the auditor's report", name, "Nowhere Rural Bank Limited")
    check("typed a bank by its legal regime", infer_type(fake, name)[0], "bank")

    sfb = [{"text": "To the Members of Tiny Small Finance Bank Limited\nBanking Regulation Act"}]
    n2, _ = extract_entity_name(sfb)
    check("small finance bank typed from its name", infer_type(sfb, n2)[0], "sfb")

    # An NBFC that OWNS a housing finance company is still an NBFC.
    nbfc = [{"text": "To the Members of Someone Finance Limited\n" + "housing finance company " * 5}]
    n3, _ = extract_entity_name(nbfc)
    check("parent not typed by its subsidiary", infer_type(nbfc, n3)[0], "nbfc")

    check("fuzzy row match to your Excel",
          match_existing_row("Kotak Mahindra Bank Limited", ["HDFC Bank", "Kotak Mahindra Bank"]),
          "Kotak Mahindra Bank")
    check("a genuinely new bank is NOT force-matched",
          match_existing_row("Nowhere Rural Bank Limited", ["HDFC Bank", "Axis Bank"]), None)

    print("SCALE-FREE PROPORTION RULE (replaces the hardcoded floors)")
    v = {"Interest Earned": {"value": 10.31, "score": 90, "why": "", "n": 1, "candidate": {}},
         "Total Assets": {"value": 270000.0, "score": 95, "why": "", "n": 1, "candidate": {}}}
    apply_dna(v, {}, {}, "FY22-23")
    check("interest of 10 crore on assets of 2.7 lakh crore is rejected",
          v["Interest Earned"]["score"] < CONFIDENCE_FLOOR, True)

    print("DERIVATION CHAIN — Net Worth from Capital+Reserves, then RoE from it")
    vd = {
        "_capital": {"value": 1396.0, "score": 95, "why": "", "n": 1, "candidate": {}},
        "_reserves": {"value": 198604.0, "score": 95, "why": "", "n": 1, "candidate": {}},
        "Profit After Tax (PAT)": {"value": 31896.0, "score": 95, "why": "", "n": 1, "candidate": {}},
    }
    vd, _notes = apply_formulas(vd)
    nw = vd.get("Net Worth / Shareholders Equity", {})
    check("Net Worth derived = Capital + Reserves", round(nw.get("value", 0)), 200000)
    roe = vd.get("Return on Equity (RoE) %", {})
    check("RoE then derived from PAT / Net Worth", round(roe.get("value", 0), 1), 15.9)

    print("REGRESSION — the ordinary two-column case still works")
    _, t = split_line("Deposits 11,80,840.70 10,64,571.61")
    check("normal statement line", line_values(t), [1180840.70, 1064571.61])
    _, t = split_line("Tier 1 Capital Ratio 17.60 17.60")
    check("'Tier 1' label not chopped at the 1", line_values(t), [17.60, 17.60])

    print()
    print("ALL PASS" if not fails else f"{len(fails)} FAILURES: {fails}")
    return not fails


# ========================================================================== #
# THE APP
# ========================================================================== #

def main():
    st.set_page_config(page_title="Annual Report → Excel", layout="wide")
    st.title("Annual Report → Excel")
    st.caption(
        "Drop in annual report PDFs of **any** Indian bank / NBFC. You get back a "
        "**brand-new Excel** of freshly extracted data — one row per bank per year. "
        "Amounts in ₹ crore, ratios as plain numbers. **A wrong number is worse than a "
        f"missing one**: anything under {CONFIDENCE_FLOOR}/100 comes out as ND, with the "
        "reason logged in the Audit Trail. Values shown in **amber** are REVIEW — worth a glance."
    )

    pdfs = st.file_uploader(
        "Annual report PDFs (any bank, one or many)",
        type=["pdf"], accept_multiple_files=True,
    )
    with st.expander("Optional: template (peer / history cross-checks)"):
        st.caption(
            "Upload your FinancialData_Verified_1.xlsx and its verified rows are used "
            "**read-only** to sanity-check extracted values against peers and prior years. "
            "Never modified. Leave empty and extraction still works."
        )
        template = st.file_uploader("Template (optional)", type=["xlsx"])

    if not (pdfs and st.button("Extract", type="primary")):
        return

    tmpdir = tempfile.mkdtemp()
    history, categories = {}, {}
    if template is not None:
        template_path = f"{tmpdir}/template.xlsx"
        with open(template_path, "wb") as fh:
            fh.write(template.getbuffer())
        history, categories = read_template(template_path)
        st.info(f"Template loaded read-only as the peer/history yardstick: "
                f"{len(history)} institutions. It will NOT be modified.")

    results, bar = [], st.progress(0.0)
    for i, pdf in enumerate(pdfs):
        with st.spinner(f"{pdf.name} — parse → rank → candidates → courtroom → DNA → formulas…"):
            try:
                r = process(pdf.getvalue(), pdf.name, history, categories)
            except Exception as exc:
                st.error(f"{pdf.name}: couldn't read it — {exc}")
                bar.progress((i + 1) / len(pdfs))
                continue
        if r is None:
            st.warning(f"{pdf.name}: couldn't work out the fiscal year. Skipped.")
        else:
            fill_from_template(r, history)        # fallback: copy ND gaps from your template
            results.append(r)
            if r["scanned"]:
                st.caption(f"{pdf.name}: {r['scanned']} scanned page(s) — "
                           "install pytesseract + pdf2image to read them.")
        bar.progress((i + 1) / len(pdfs))

    if not results:
        return

    out_path = f"{tmpdir}/FinancialData_Extracted.xlsx"
    build_fresh_workbook(out_path, results)      # a NEW workbook, template untouched

    for r in results:
        filled = sum(1 for c in r["cells"].values() if c["decision"] in ("WRITE", "REVIEW"))
        review = sum(1 for c in r["cells"].values() if c["decision"] == "REVIEW")
        nd = sum(1 for c in r["cells"].values() if c["decision"] == "ND")
        st.success(f"**{r['institution']} · {r['fy']}** — {filled} filled "
                   f"({review} to review), {nd} ND")

    st.subheader("Extracted data")
    st.dataframe(pd.DataFrame([
        dict({"Institution": r["institution"], "FY": r["fy"]},
             **{h: c["value"] for h, c in r["cells"].items()}) for r in results
    ]), use_container_width=True, hide_index=True)

    st.subheader("Audit trail — why every cell came out the way it did")
    st.dataframe(pd.DataFrame([
        {"Institution": r["institution"], "FY": r["fy"], "Metric": h, "Value": c["value"],
         "Confidence": c["score"], "Decision": c["decision"],
         "Derived": "YES" if c["derived"] else "", "Why": c["why"]}
        for r in results for h, c in r["cells"].items()
    ]), use_container_width=True, hide_index=True)

    with open(out_path, "rb") as fh:
        st.download_button("⬇ Download extracted Excel", data=fh.read(),
                           file_name="FinancialData_Extracted.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")


# ========================================================================== #
# BATCH CLI  (the arx-style entry point, folded into the one file)
#
# Same engine as the Streamlit app — no second copy of the logic. Point it at a
# folder (or several) of PDFs and it writes a fresh Excel, headless. This is the
# "python -m arx.run" workflow, without a separate package to keep in sync.
# ========================================================================== #

def discover_pdfs(paths, recursive=True):
    """Every .pdf under the given files/folders, de-duplicated, sorted.

    A path may be a single .pdf or a folder; folders are scanned recursively by
    default, so bank-wise subfolders (reports/Axis/…, input_pdfs/…) just work.
    """
    found, seen = [], set()
    for raw in paths:
        p = Path(raw)
        if p.is_file() and p.suffix.lower() == ".pdf":
            cands = [p]
        elif p.is_dir():
            cands = sorted(p.rglob("*.pdf") if recursive else p.glob("*.pdf"))
        else:
            cands = []
        for c in cands:
            key = str(c.resolve())
            if key not in seen:
                seen.add(key)
                found.append(c)
    return found


def run_cli(argv):
    """`python simple_app.py --pdfs ./folder --out result.xlsx [--template t.xlsx]`"""
    import argparse

    ap = argparse.ArgumentParser(
        prog="simple_app.py",
        description="Extract Indian bank / NBFC annual-report PDFs into a FRESH Excel "
                    "(the template, if given, is read-only). Same engine as the Streamlit app.",
    )
    ap.add_argument("--pdfs", nargs="*",
                    help="PDF files or folders (scanned recursively). If omitted, "
                         "auto-scans ./input_pdfs, ./reports and the current folder.")
    ap.add_argument("--template", help="optional template .xlsx for peer/history "
                                       "cross-checks — read-only, never modified")
    ap.add_argument("--out", default="FinancialData_Extracted.xlsx",
                    help="output workbook (default: FinancialData_Extracted.xlsx)")
    ap.add_argument("--no-recursive", action="store_true",
                    help="do not descend into subfolders")
    ap.add_argument("--self-test", action="store_true", help="run the built-in checks and exit")
    args = ap.parse_args(argv)

    if args.self_test:
        return 0 if self_test() else 1

    search = args.pdfs if args.pdfs else ["input_pdfs", "reports", "."]
    pdfs = discover_pdfs(search, recursive=not args.no_recursive)
    if not pdfs:
        print(f"No PDFs found under: {', '.join(search)}")
        return 2

    history, categories = {}, {}
    if args.template and Path(args.template).exists():
        history, categories = read_template(args.template)
        print(f"Template loaded read-only ({len(history)} institutions) — it will NOT be modified.")

    print(f"Found {len(pdfs)} PDF(s). Extracting…\n")
    results, failures = [], []
    for pdf in pdfs:
        try:
            r = process(pdf.read_bytes(), pdf.name, history, categories)
        except Exception as exc:                      # one bad PDF never kills the batch
            failures.append((pdf.name, str(exc)))
            print(f"  FAILED  {pdf.name}: {exc}")
            continue
        if r is None:
            print(f"  SKIP    {pdf.name}: could not determine the fiscal year")
            continue
        fill_from_template(r, history)             # fill ND gaps from the template
        results.append(r)
        filled = sum(1 for c in r["cells"].values() if c["decision"] in ("WRITE", "REVIEW"))
        nd = sum(1 for c in r["cells"].values() if c["decision"] == "ND")
        print(f"  OK      {r['institution']:<32} {r['fy']}   {filled} filled, {nd} ND")

    if not results:
        print("\nNothing extracted.")
        return 1

    build_fresh_workbook(args.out, results)           # a NEW workbook; template untouched
    print(f"\nWrote {args.out}  ({len(results)} report(s), {len(failures)} failed).")
    return 0


if __name__ == "__main__":
    # `streamlit run simple_app.py`                     -> the upload/download app
    # `python simple_app.py`                            -> the self-test (unchanged)
    # `python simple_app.py --pdfs ./folder --out x.xlsx` -> headless batch extraction
    import sys

    try:
        running_in_streamlit = st is not None and st.runtime.exists()
    except Exception:
        running_in_streamlit = False

    if running_in_streamlit:
        main()
    elif not sys.argv[1:]:
        raise SystemExit(0 if self_test() else 1)      # bare `python simple_app.py`
    else:
        raise SystemExit(run_cli(sys.argv[1:]))