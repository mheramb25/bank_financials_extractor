"""The golden test.

Two levels:

1. **Synthetic golden (always runs).**  ``tests/conftest.py`` builds a Document
   with the structure of ICICI Bank's FY22-23 report -- two-column statements,
   Indian grouping, a unit caption, Highlights, Basel III, and a narrative decoy.
   The whole Stage 3-16 chain runs over it and must reproduce the template's ICICI
   row.  This runs in milliseconds, offline, with no PDF, no OCR, no camelot.

2. **Real-PDF golden (runs when the file is present).**  Drop the actual report at
   ``tests/fixtures/icici_fy2223.pdf`` and the same assertions run against the
   real thing, through the real parser.  **The template's ICICI row is the ground
   truth and the regression fixture.**

Both assert the same numbers:
    Total Assets 1584207 | Deposits 1180841 | Gross Advances 1019638
    Interest Earned 109231.34 | Interest Expended 47102.74 | NII 62129
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from arx.confidence import Decision
from arx.excel_writer import write_results
from arx.extract import collect_evidence, compile_metrics, generate_candidates
from arx.pipeline import build_document, process_document
from arx.rank import rank_pages

from .conftest import ICICI_FY2223_TRUTH

REAL_PDF = Path(__file__).parent / "fixtures" / "icici_fy2223.pdf"

# Amounts: +/-1% relative. Ratios: +/-0.1 percentage points.
AMOUNT_REL = 0.01
RATIO_ABS = 0.1
RATIOS = {"gnpa_ratio", "crar"}


def assert_matches_truth(result, keys=tuple(ICICI_FY2223_TRUTH)):
    values = result.values()
    for key in keys:
        expected = ICICI_FY2223_TRUTH[key]
        cell = result.cell(key)
        assert cell is not None, f"{key}: no cell produced at all"
        assert cell.sentinel is None, (
            f"{key}: came out as {cell.sentinel} -- {cell.reason}"
        )
        got = values[key]
        if key in RATIOS:
            assert got == pytest.approx(expected, abs=RATIO_ABS), f"{key}: {got} != {expected}"
        else:
            assert got == pytest.approx(expected, rel=AMOUNT_REL), f"{key}: {got} != {expected}"


# --------------------------------------------------------------------------- #
# Stage 2: identification
# --------------------------------------------------------------------------- #


class TestIdentification:
    def test_institution_is_recognised(self, icici_document):
        assert icici_document.institution == "ICICI Bank Limited"
        assert icici_document.category == "D-SIB"
        assert icici_document.inst_type == "bank"

    def test_fy_ending_march_2023_is_fy22_23(self, icici_document):
        assert icici_document.fiscal_year == "FY22-23"
        assert icici_document.prior_fiscal_year == "FY21-22"


# --------------------------------------------------------------------------- #
# Stage 3: ranking
# --------------------------------------------------------------------------- #


class TestRanking:
    def test_statements_outrank_the_chairmans_letter(self, icici_document):
        ranked = rank_pages(icici_document)
        order = [p.section.value for p in ranked]
        assert order[0] == "financial_statements"
        assert order[-1] == "narrative"


# --------------------------------------------------------------------------- #
# Stage 4-16: the whole chain, on the synthetic report
# --------------------------------------------------------------------------- #


class TestSyntheticGolden:
    @pytest.fixture()
    def result(self, icici_document):
        pool = collect_evidence(
            generate_candidates(icici_document, rank_pages(icici_document), compile_metrics())
        )
        return process_document(icici_document, pool)

    def test_extracted_values_match_the_template_row(self, result):
        assert_matches_truth(result)

    def test_the_prior_year_column_was_read_but_not_written(self, result):
        # FY21-22 numbers exist in the report and we captured them as a
        # cross-check -- but the FY22-23 cells hold FY22-23 numbers.
        assert result.prior_year_values["total_assets"] == pytest.approx(
            1411297.74, rel=0.001
        )
        assert result.values()["total_assets"] == pytest.approx(1584207.0, rel=AMOUNT_REL)

    def test_headline_metrics_clear_the_write_threshold(self, result):
        for key in ("total_assets", "total_deposits", "gross_advances", "pat", "nii"):
            cell = result.cell(key)
            assert cell.decision in (Decision.AUTO, Decision.HIGH, Decision.MANUAL)
            assert cell.confidence >= 65

    def test_metrics_that_do_not_apply_to_a_bank_are_NA_not_ND(self, result):
        for key in ("aum", "stage3_pct"):
            cell = result.cell(key)
            assert cell.sentinel == "NA"
            assert cell.decision == Decision.NOT_APPLICABLE
            assert "not applicable" in cell.reason.lower()

    def test_undisclosed_metrics_are_ND_with_a_reason(self, result):
        cell = result.cell("esg_rating")
        assert cell.sentinel == "ND"
        assert cell.reason.strip()

    def test_the_narrative_decoy_did_not_win(self, result):
        # p7 says "profit after tax crossed 30,000 crore". The P&L says 31,896.50.
        assert result.values()["pat"] == pytest.approx(31896.50, rel=AMOUNT_REL)
        assert result.cell("pat").section == "financial_statements"

    def test_formula_identities_hold_on_the_chosen_values(self, result):
        v = result.values()
        assert v["nii"] == pytest.approx(
            v["interest_earned"] - v["interest_expended"], rel=AMOUNT_REL
        )
        assert v["crar"] == pytest.approx(v["tier1"] + v["tier2"], abs=RATIO_ABS)

    def test_every_cell_carries_an_audit_reason(self, result):
        for cell in result.cells:
            assert cell.reason.strip(), f"{cell.metric} has no reason"

    def test_the_run_is_deterministic(self, icici_document):
        def once():
            pool = collect_evidence(
                generate_candidates(
                    icici_document, rank_pages(icici_document), compile_metrics()
                )
            )
            return process_document(icici_document, pool).values()

        assert once() == once()


# --------------------------------------------------------------------------- #
# Stage 16: all the way into the workbook
# --------------------------------------------------------------------------- #


class TestWorkbookOutput:
    def test_the_filled_workbook_holds_the_template_row(
        self, icici_document, blank_template, tmp_path
    ):
        pool = collect_evidence(
            generate_candidates(icici_document, rank_pages(icici_document), compile_metrics())
        )
        result = process_document(icici_document, pool)
        out = write_results([result], blank_template, tmp_path / "FinancialData_Filled.xlsx")

        ws = load_workbook(out)["FY22-23"]
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=3, column=3).value == "ICICI Bank Limited"
        assert ws.cell(row=3, column=headers["Total Assets"]).value == pytest.approx(
            1584207.0, rel=AMOUNT_REL
        )
        assert ws.cell(row=3, column=headers["Interest Earned"]).value == pytest.approx(
            109231.34, rel=AMOUNT_REL
        )
        # Nothing was written as a string with a currency symbol.
        for header in ("Total Assets", "Total Deposits", "Net Interest Income (NII)"):
            assert isinstance(ws.cell(row=3, column=headers[header]).value, (int, float))


# --------------------------------------------------------------------------- #
# The real PDF, when you have it
# --------------------------------------------------------------------------- #


@pytest.mark.skipif(
    not REAL_PDF.exists(),
    reason=f"drop the real annual report at {REAL_PDF} to run the real-PDF golden test",
)
class TestRealPdfGolden:
    @pytest.fixture(scope="class")
    def result(self):
        doc, pool = build_document(REAL_PDF)
        return process_document(doc, pool)

    def test_identification(self, result):
        assert result.institution == "ICICI Bank Limited"
        assert result.fiscal_year == "FY22-23"

    def test_headline_numbers_match_the_template(self, result):
        assert_matches_truth(
            result,
            keys=(
                "total_assets",
                "total_deposits",
                "gross_advances",
                "interest_earned",
                "interest_expended",
                "nii",
            ),
        )

    def test_asset_quality_and_capital_match(self, result):
        assert_matches_truth(result, keys=("gnpa_ratio", "crar"))
