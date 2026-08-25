"""Stage 16: the writer must not damage the template, and must not lie."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from arx import load_config, load_metrics
from arx.excel_writer import TemplateWriter, read_history, write_results
from arx.models import CellResult, Decision, DocumentResult, ProsecutorHit

CFG = load_config()
ND = CFG["excel"]["nd_sentinel"]
NA = CFG["excel"]["na_sentinel"]


def cell(metric, header, value=None, sentinel=None, confidence=92.0, decision=Decision.HIGH, reason="ok"):
    return CellResult(
        institution="ICICI Bank Limited",
        fiscal_year="FY22-23",
        metric=metric,
        excel_header=header,
        value=value,
        sentinel=sentinel,
        confidence=confidence,
        decision=decision,
        reason=reason,
        pages=[120],
        section="financial_statements",
        alias_matched=r"Total\s+Assets",
        unit_as_printed="crore",
        candidate_count=4,
        defence_score=98.0,
        prosecutor_hits=[],
        checks_passed=["range:total_assets"],
        checks_failed=[],
        rejected=["1411297.74@p120(38)"],
    )


def icici_result(**overrides) -> DocumentResult:
    cells = [
        cell("total_assets", "Total Assets", 1584207.0),
        cell("total_deposits", "Total Deposits", 1180841.0),
        cell("gross_advances", "Total Advances (Gross)", 1019638.0),
        cell(
            "aum",
            "Assets Under Management (AUM)",
            sentinel=NA,
            decision=Decision.NOT_APPLICABLE,
            confidence=100.0,
            reason="Metric not applicable to a bank (D-SIB)",
        ),
        cell(
            "esg_rating",
            "ESG Rating (CRISIL/MSCI)",
            sentinel=ND,
            decision=Decision.REJECT,
            confidence=41.0,
            reason="Found only in narrative text (confidence 41 < 65 threshold)",
        ),
        cell("gnpa_ratio", "Gross NPA Ratio %", 2.81, confidence=71.0, decision=Decision.MANUAL),
    ]
    base = dict(
        path="icici.pdf",
        institution="ICICI Bank Limited",
        category="D-SIB",
        inst_type="bank",
        fiscal_year="FY22-23",
        prior_fiscal_year="FY21-22",
        cells=cells,
    )
    base.update(overrides)
    return DocumentResult(**base)


class TestTemplatePreservation:
    def test_sheet_order_and_headers_survive(self, blank_template, tmp_path):
        before = load_workbook(blank_template)
        before_sheets = list(before.sheetnames)
        before_headers = [c.value for c in before["FY22-23"][2]]
        before_width = before["FY22-23"].column_dimensions["C"].width

        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)

        after = load_workbook(out)
        # Original sheets are all still there, in order, before the new ones.
        assert after.sheetnames[: len(before_sheets)] == before_sheets
        assert [c.value for c in after["FY22-23"][2]] == before_headers
        assert after["FY22-23"].column_dimensions["C"].width == before_width
        assert "Source References" in after.sheetnames

    def test_the_template_file_itself_is_never_touched(self, blank_template, tmp_path):
        stamp = blank_template.stat().st_mtime_ns
        digest_before = blank_template.read_bytes()
        write_results([icici_result()], blank_template, tmp_path / "filled.xlsx")
        assert blank_template.read_bytes() == digest_before
        assert blank_template.stat().st_mtime_ns == stamp

    def test_audit_sheets_are_added(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        wb = load_workbook(out)
        assert "Audit Trail" in wb.sheetnames
        assert "Confidence Summary" in wb.sheetnames
        assert "Missing Cells Report" in wb.sheetnames


class TestValues:
    def test_numbers_are_written_as_numbers(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["FY22-23"]
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(row=3, column=headers["Total Assets"]).value
        assert isinstance(val, (int, float))
        assert val == pytest.approx(1584207.0)
        # ...and ratios are plain numbers: 2.81 means 2.81%, not 0.0281.
        assert ws.cell(row=3, column=headers["Gross NPA Ratio %"]).value == pytest.approx(2.81)

    def test_nd_and_na_sentinels_are_written(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["FY22-23"]
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=3, column=headers["Assets Under Management (AUM)"]).value == NA
        assert ws.cell(row=3, column=headers["ESG Rating (CRISIL/MSCI)"]).value == ND

    def test_low_confidence_cells_get_a_comment(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["FY22-23"]
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        c = ws.cell(row=3, column=headers["Gross NPA Ratio %"])
        assert c.comment is not None
        assert "71" in c.comment.text


class TestRows:
    def test_new_institution_is_appended_with_the_next_sr_no(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["FY22-23"]
        assert ws.cell(row=3, column=1).value == 1
        assert ws.cell(row=3, column=2).value == "D-SIB"
        assert ws.cell(row=3, column=3).value == "ICICI Bank Limited"

    def test_existing_row_is_updated_in_place_and_matched_fuzzily(
        self, blank_template, tmp_path
    ):
        # Seed the template with a row whose name is spelled differently.
        wb = load_workbook(blank_template)
        ws = wb["FY22-23"]
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value="D-SIB")
        ws.cell(row=3, column=3, value="ICICI Bank Ltd.")
        ws.cell(row=4, column=1, value=2)
        ws.cell(row=4, column=2, value="D-SIB")
        ws.cell(row=4, column=3, value="HDFC Bank Limited")
        wb.save(blank_template)

        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws2 = load_workbook(out)["FY22-23"]

        # Updated in place: no new row, HDFC untouched, name spelling preserved.
        assert ws2.cell(row=3, column=3).value == "ICICI Bank Ltd."
        assert ws2.cell(row=4, column=3).value == "HDFC Bank Limited"
        assert ws2.cell(row=5, column=3).value is None
        headers = {ws2.cell(row=2, column=c).value: c for c in range(1, ws2.max_column + 1)}
        assert ws2.cell(row=3, column=headers["Total Assets"]).value == pytest.approx(1584207.0)

    def test_sbi_alias_matches_the_template_spelling(self, blank_template, tmp_path):
        wb = load_workbook(blank_template)
        ws = wb["FY22-23"]
        ws.cell(row=3, column=3, value="State Bank of India (SBI)")
        wb.save(blank_template)

        result = icici_result(
            institution="State Bank of India (SBI)", category="D-SIB"
        )
        for c in result.cells:
            c.institution = "State Bank of India (SBI)"
        out = tmp_path / "filled.xlsx"
        write_results([result], blank_template, out)
        ws2 = load_workbook(out)["FY22-23"]
        assert ws2.cell(row=4, column=3).value is None  # not appended: matched row 3


class TestMissingCellsReport:
    def test_every_nd_and_na_gets_a_row_with_a_real_reason(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["Missing Cells Report"]
        rows = [
            [c.value for c in row]
            for row in ws.iter_rows(min_row=2)
            if row[0].value
        ]
        metrics = {r[2] for r in rows}
        assert "Assets Under Management (AUM)" in metrics
        assert "ESG Rating (CRISIL/MSCI)" in metrics
        for r in rows:
            assert r[3] and len(str(r[3])) > 10  # a real machine-generated reason

    def test_headers_are_exactly_the_agreed_four(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["Missing Cells Report"]
        assert [c.value for c in ws[1]][:4] == [
            "Institution",
            "Fiscal Year",
            "Metric",
            "Reason data unavailable",
        ]


class TestIdempotenceAndSafety:
    def test_nd_never_overwrites_a_real_number(self, blank_template, tmp_path):
        # A previous run wrote a good Total Assets; this run cannot find it.
        wb = load_workbook(blank_template)
        ws = wb["FY22-23"]
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=3, value="ICICI Bank Limited")
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        ws.cell(row=3, column=headers["Total Assets"], value=1584207.0)
        wb.save(blank_template)

        blind = icici_result(
            cells=[
                cell(
                    "total_assets",
                    "Total Assets",
                    sentinel=ND,
                    decision=Decision.REJECT,
                    confidence=12.0,
                    reason="no candidate found",
                )
            ]
        )
        out = tmp_path / "filled.xlsx"
        write_results([blind], blank_template, out)
        ws2 = load_workbook(out)["FY22-23"]
        assert ws2.cell(row=3, column=headers["Total Assets"]).value == pytest.approx(
            1584207.0
        )

    def test_a_cell_only_upgrades_on_higher_confidence(self, blank_template, tmp_path):
        writer = TemplateWriter(blank_template, tmp_path / "filled.xlsx")
        strong = icici_result(cells=[cell("total_assets", "Total Assets", 1584207.0, confidence=93.0)])
        weak = icici_result(cells=[cell("total_assets", "Total Assets", 999999.0, confidence=70.0, decision=Decision.MANUAL)])
        writer.write_result(strong)
        writer.write_result(weak)
        out = writer.save()

        ws = load_workbook(out)["FY22-23"]
        headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=3, column=headers["Total Assets"]).value == pytest.approx(1584207.0)

        audit = load_workbook(out)["Audit Trail"]
        reasons = [row[15].value or "" for row in audit.iter_rows(min_row=2)]
        assert any("kept existing value" in str(r) for r in reasons)

    def test_a_missing_fy_sheet_is_cloned_from_the_donor_layout(
        self, blank_template, tmp_path
    ):
        result = icici_result(fiscal_year="FY24-25")
        for c in result.cells:
            c.fiscal_year = "FY24-25"
        out = tmp_path / "filled.xlsx"
        write_results([result], blank_template, out)

        wb = load_workbook(out)
        assert "FY24-25" in wb.sheetnames
        assert [c.value for c in wb["FY24-25"][2]] == [c.value for c in wb["FY23-24"][2]]
        assert "FY24-25" in str(wb["FY24-25"].cell(row=1, column=1).value)
        # And it starts empty except for our institution.
        assert wb["FY24-25"].cell(row=3, column=3).value == "ICICI Bank Limited"


class TestAuditTrail:
    def test_one_row_per_extracted_cell_with_all_the_columns(
        self, blank_template, tmp_path
    ):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["Audit Trail"]
        headers = [c.value for c in ws[1]]
        for required in (
            "Chosen Value",
            "Page(s)",
            "Defence score",
            "Prosecutor penalties",
            "Final confidence",
            "Decision (Auto/High/Manual/Reject)",
            "Rejected candidates (value@page, ...)",
        ):
            assert required in headers
        body = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(body) == 6  # one per cell in the result

    def test_confidence_summary_counts_the_bands(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        ws = load_workbook(out)["Confidence Summary"]
        row = next(r for r in ws.iter_rows(min_row=2, values_only=True) if r[0])
        institution, fy, filled, nd, na = row[0], row[1], row[2], row[3], row[4]
        assert institution == "ICICI Bank Limited"
        assert fy == "FY22-23"
        assert filled == 4
        assert nd == 1
        assert na == 1
        assert "Gross NPA Ratio %" in str(row[10])  # the Manual-Review cell list


class TestReadHistory:
    def test_history_round_trips(self, blank_template, tmp_path):
        out = tmp_path / "filled.xlsx"
        write_results([icici_result()], blank_template, out)
        history = read_history(out)
        assert history["ICICI Bank Limited"]["total_assets"]["FY22-23"] == pytest.approx(
            1584207.0
        )
